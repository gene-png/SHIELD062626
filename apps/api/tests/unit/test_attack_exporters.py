"""ATT&CK PDF + XLSX exporter smokes."""

from __future__ import annotations

import io
import uuid

import pytest

from app.attack.analytics import compute as compute_heatmap
from app.attack.catalog import TECHNIQUES
from app.attack.coverage import CoverageStatus
from app.attack.exporters import build_context, render_docx, render_pdf, render_xlsx
from app.models.attack_assessment import (
    AttackAssessment,
    AttackAssessmentStatus,
    AttackCoverage,
)


def _build_inputs(*, default_status: str | None = "covered"):
    a = AttackAssessment(
        id=uuid.uuid4(),
        service_id=uuid.uuid4(),
        version=1,
        status=AttackAssessmentStatus.APPROVED,
    )
    coverage: list[AttackCoverage] = []
    for t in TECHNIQUES:
        coverage.append(
            AttackCoverage(
                id=uuid.uuid4(),
                assessment_id=a.id,
                technique_code=t.id,
                status=default_status,
            )
        )
    coverage_map = {c.technique_code: c.status for c in coverage}
    rollup = compute_heatmap(coverage_map)
    return a, coverage, rollup


def _ctx(*, default_status: str | None = "covered"):
    a, coverage, rollup = _build_inputs(default_status=default_status)
    return build_context(
        client_legal_name="Atlas Defense Solutions",
        service_title="MITRE ATT&CK Coverage",
        assessment=a,
        coverage=coverage,
        rollup=rollup,
    )


@pytest.mark.unit
def test_xlsx_has_three_sheets() -> None:
    from openpyxl import load_workbook

    raw = render_xlsx(_ctx())
    assert raw[:2] == b"PK"
    wb = load_workbook(io.BytesIO(raw))
    assert set(wb.sheetnames) == {"Heatmap Summary", "Coverage", "Gaps"}


@pytest.mark.unit
def test_xlsx_coverage_sheet_has_one_row_per_technique() -> None:
    from openpyxl import load_workbook

    raw = render_xlsx(_ctx())
    wb = load_workbook(io.BytesIO(raw))
    ws = wb["Coverage"]
    assert ws.max_row == len(TECHNIQUES) + 1


@pytest.mark.unit
def test_xlsx_gap_sheet_lists_only_gaps() -> None:
    from openpyxl import load_workbook

    ctx = _ctx(default_status=CoverageStatus.GAP.value)
    raw = render_xlsx(ctx)
    wb = load_workbook(io.BytesIO(raw))
    ws = wb["Gaps"]
    # Header + every technique.
    assert ws.max_row == len(TECHNIQUES) + 1


@pytest.mark.unit
def test_xlsx_gap_placeholder_when_no_gaps() -> None:
    from openpyxl import load_workbook

    ctx = _ctx(default_status=CoverageStatus.COVERED.value)
    raw = render_xlsx(ctx)
    wb = load_workbook(io.BytesIO(raw))
    ws = wb["Gaps"]
    assert ws.max_row == 2  # header + single placeholder
    assert ws.cell(row=2, column=2).value == "No gaps recorded"


@pytest.mark.unit
def test_pdf_renders_valid_bytes() -> None:
    raw = render_pdf(_ctx())
    assert raw.startswith(b"%PDF-")
    assert len(raw) > 2000


def _pdf_text(raw: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(raw))
    return "".join(page.extract_text() for page in reader.pages)


@pytest.mark.unit
def test_pdf_carries_title_client_and_a_known_tactic() -> None:
    # SMOKE §10 content assertions. The old version matched only "MITRE" with a
    # comment normalizing reportlab's "ATT&CK;" entity artifact — it encoded
    # the very header bug the fix removes (a bare "&" fed to Paragraph markup,
    # which reportlab re-emits as an unknown entity with a synthesized
    # semicolon). The exact title must render, ampersand intact.
    text = _pdf_text(render_pdf(_ctx()))
    assert "MITRE ATT&CK Coverage" in text  # exact service title
    assert "ATT&CK;" not in text  # the entity artifact must be gone
    assert "Atlas Defense Solutions" in text  # client name
    assert "Reconnaissance" in text  # a known tactic rollup row


@pytest.mark.unit
def test_pdf_escapes_ampersand_in_client_name() -> None:
    # A client whose legal name carries "&" must render literally, not as a
    # mangled entity (same escape path as the title line).
    a, coverage, rollup = _build_inputs()
    ctx = build_context(
        client_legal_name="Rook&Pawn Security",
        service_title="MITRE ATT&CK Coverage",
        assessment=a,
        coverage=coverage,
        rollup=rollup,
    )
    text = _pdf_text(render_pdf(ctx))
    assert "Rook&Pawn Security" in text


@pytest.mark.unit
def test_pdf_handles_zero_gaps() -> None:
    ctx = _ctx(default_status=CoverageStatus.COVERED.value)
    raw = render_pdf(ctx)
    assert raw.startswith(b"%PDF-")


@pytest.mark.unit
def test_build_context_falls_back_when_client_none() -> None:
    a, coverage, rollup = _build_inputs()
    ctx = build_context(
        client_legal_name=None,
        service_title="x",
        assessment=a,
        coverage=coverage,
        rollup=rollup,
    )
    assert ctx.client_legal_name == "Client"


# ---------------------------------------------------------------------------
# S2 (D-035): curated evidence columns, citation-only gap labels, tactic heatmap
# ---------------------------------------------------------------------------

# The first two catalog codes, used to plant a curated row and a bare row.
_FIRST_CODE = TECHNIQUES[0].id
_SECOND_CODE = TECHNIQUES[1].id


def _ctx_with(
    *,
    default_status: str | None = "covered",
    curated: dict[str, dict] | None = None,
    evidence_names: dict | None = None,
):
    """A context where named technique codes carry curated tools/evidence."""
    a, coverage, _ = _build_inputs(default_status=default_status)
    by_code = {c.technique_code: c for c in coverage}
    for code, fields in (curated or {}).items():
        row = by_code[code]
        for key, value in fields.items():
            setattr(row, key, value)
    rollup = compute_heatmap({c.technique_code: c.status for c in coverage})
    return build_context(
        client_legal_name="Atlas Defense Solutions",
        service_title="MITRE ATT&CK Coverage",
        assessment=a,
        coverage=coverage,
        rollup=rollup,
        evidence_names=evidence_names,
    )


def _coverage_row(ws, code: str) -> tuple:
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == code:
            return row
    raise AssertionError(f"technique {code} missing from the Coverage sheet")


@pytest.mark.unit
def test_xlsx_coverage_sheet_header_contract() -> None:
    """The Coverage sheet's column contract, in order. Curated tool citations,
    rationale and the evidence reference are client-visible columns now."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(render_xlsx(_ctx())))
    headers = [c.value for c in wb["Coverage"][1]]
    assert headers == [
        "Technique",
        "Name",
        "Tactic(s)",
        "Type",
        "Status",
        "Detection tools",
        "Prevention tools",
        "Response tools",
        "Rationale",
        "Evidence reference",
        "Notes",
    ]


@pytest.mark.unit
def test_xlsx_coverage_joins_tool_lists_with_comma_space() -> None:
    from openpyxl import load_workbook

    ctx = _ctx_with(
        curated={
            _FIRST_CODE: {
                "detection_tools": ["Okta", "CrowdStrike"],
                "prevention_tools": ["Okta"],
                "response_tools": [],
                "rationale": "Sign-in telemetry is forwarded to the SIEM.",
            }
        }
    )
    wb = load_workbook(io.BytesIO(render_xlsx(ctx)))
    row = _coverage_row(wb["Coverage"], _FIRST_CODE)
    assert row[5] == "Okta, CrowdStrike"  # Detection tools, joined
    assert row[6] == "Okta"
    # An empty list renders blank, never "[]". openpyxl reads a blank cell as None.
    assert row[7] in (None, "")
    assert row[8] == "Sign-in telemetry is forwarded to the SIEM."


@pytest.mark.unit
def test_xlsx_evidence_reference_resolves_attached_filename() -> None:
    from openpyxl import load_workbook

    art_id = uuid.uuid4()
    ctx = _ctx_with(
        curated={_FIRST_CODE: {"evidence_artifact_id": art_id}},
        evidence_names={art_id: "atlas-siem-export-2026-07.csv"},
    )
    wb = load_workbook(io.BytesIO(render_xlsx(ctx)))
    assert _coverage_row(wb["Coverage"], _FIRST_CODE)[9] == "atlas-siem-export-2026-07.csv"


@pytest.mark.unit
def test_xlsx_evidence_reference_says_none_attached_when_null() -> None:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(render_xlsx(_ctx())))
    assert _coverage_row(wb["Coverage"], _SECOND_CODE)[9] == "No evidence attached"


@pytest.mark.unit
def test_build_context_raises_on_an_unresolvable_evidence_reference() -> None:
    """FAIL LOUDLY: a dangling evidence pointer must not print as
    "No evidence attached" — that sentence means the column is genuinely NULL."""
    art_id = uuid.uuid4()
    with pytest.raises(ValueError, match="evidence_names"):
        _ctx_with(
            curated={_FIRST_CODE: {"evidence_artifact_id": art_id}},
            evidence_names={},
        )


@pytest.mark.unit
def test_xlsx_gaps_sheet_header_contract() -> None:
    from openpyxl import load_workbook

    ctx = _ctx_with(default_status=CoverageStatus.GAP.value)
    wb = load_workbook(io.BytesIO(render_xlsx(ctx)))
    assert [c.value for c in wb["Gaps"][1]] == [
        "Technique",
        "Name",
        "Tactic(s)",
        "Gap Direction",
        "Notes",
    ]


@pytest.mark.unit
def test_gap_direction_states_the_absence_of_a_citation_verbatim() -> None:
    from openpyxl import load_workbook

    ctx = _ctx_with(default_status=CoverageStatus.GAP.value)
    wb = load_workbook(io.BytesIO(render_xlsx(ctx)))
    row = _coverage_row(wb["Gaps"], _FIRST_CODE)
    assert row[3] == "No detection, prevention, or response tool is cited for this technique"


@pytest.mark.unit
def test_gap_direction_names_the_cited_tools_when_some_exist() -> None:
    from openpyxl import load_workbook

    ctx = _ctx_with(
        default_status=CoverageStatus.GAP.value,
        curated={
            _FIRST_CODE: {
                "detection_tools": ["Okta"],
                "response_tools": ["CrowdStrike"],
            }
        },
    )
    wb = load_workbook(io.BytesIO(render_xlsx(ctx)))
    assert _coverage_row(wb["Gaps"], _FIRST_CODE)[3] == "Cited: Okta, CrowdStrike (partial)"


@pytest.mark.unit
def test_gap_direction_never_states_cause_or_remediation() -> None:
    """D-035 label discipline. Run AI overwrites tools and rationale on every
    unlocked row, so a Gap Direction cell may report only what is cited. Any
    causal or remedial vocabulary in this column is a product defect."""
    from openpyxl import load_workbook

    forbidden = (
        "because",
        "consider",
        "deploy",
        "implement",
        "lack",
        "mitigat",
        "recommend",
        "remediat",
        "risk",
        "should",
        "unprotected",
        "vulnerab",
    )
    ctx = _ctx_with(
        default_status=CoverageStatus.GAP.value,
        curated={_FIRST_CODE: {"detection_tools": ["Okta"]}},
    )
    wb = load_workbook(io.BytesIO(render_xlsx(ctx)))
    cells = [r[3] for r in wb["Gaps"].iter_rows(min_row=2, values_only=True)]
    assert cells, "the Gaps sheet rendered no Gap Direction cells to audit"
    for cell in cells:
        lowered = (cell or "").lower()
        for word in forbidden:
            assert word not in lowered, f"Gap Direction leaked inference: {cell!r}"


@pytest.mark.unit
def test_heatmap_summary_coverage_pct_cells_carry_coverage_hex_fills() -> None:
    from openpyxl import load_workbook

    from app.attack.exporters import coverage_hex

    ctx = _ctx()
    wb = load_workbook(io.BytesIO(render_xlsx(ctx)))
    ws = wb["Heatmap Summary"]

    def argb(pct: float) -> str:
        return "FF" + coverage_hex(pct)[1:].upper()

    # The overall figure on the header block.
    assert ws.cell(row=4, column=2).fill.start_color.rgb == argb(ctx.rollup.coverage_pct)
    # Every per-tactic row. Header block is rows 1-5, blank 6, table header 7.
    for offset, tc in enumerate(ctx.rollup.by_tactic):
        cell = ws.cell(row=8 + offset, column=10)
        assert cell.value == tc.coverage_pct
        assert cell.fill.start_color.rgb == argb(tc.coverage_pct)


@pytest.mark.unit
def test_coverage_hex_raises_outside_zero_to_one_hundred() -> None:
    """S1's ramp raises rather than clamping; the coverage band inherits that."""
    from app.attack.exporters import coverage_hex

    assert coverage_hex(0.0) != coverage_hex(100.0)
    for bad in (-0.1, 100.1, 250.0):
        with pytest.raises(ValueError, match="coverage_pct"):
            coverage_hex(bad)


def _pdf_norm(raw: bytes) -> str:
    """Extracted PDF text with reportlab's line wraps collapsed to spaces."""
    return " ".join(_pdf_text(raw).split())


_METHODOLOGY_CLAUSES = (
    "drafted by Run AI",
    "A consultant can edit any of these fields",
    "locking a row keeps a later Run AI run from overwriting it",
    "no field here should be read as verified",
    "Substantiation states arrive in the next batch of work",
)


@pytest.mark.unit
def test_pdf_carries_the_citation_defensibility_stat() -> None:
    total = len(TECHNIQUES)
    # No row carries a tool, so the honest figure is zero.
    assert f"0 of {total} scored techniques cite at least one tool" in _pdf_norm(render_pdf(_ctx()))
    # Two curated rows move the numerator, computed in Python from the rows.
    ctx = _ctx_with(
        curated={
            _FIRST_CODE: {"detection_tools": ["Okta"]},
            _SECOND_CODE: {"response_tools": ["CrowdStrike"]},
        }
    )
    assert f"2 of {total} scored techniques cite at least one tool" in _pdf_norm(render_pdf(ctx))


@pytest.mark.unit
def test_pdf_carries_the_methodology_disclosure() -> None:
    text = _pdf_norm(render_pdf(_ctx()))
    for clause in _METHODOLOGY_CLAUSES:
        assert clause in text, f"methodology note lost: {clause!r}"


@pytest.mark.unit
def test_docx_carries_the_citation_stat_and_the_methodology_disclosure() -> None:
    """The criterion names the PDF and the DOCX but its evidence clause only
    reached the PDF; python-docx paragraphs are exact, so assert them too."""
    from docx import Document

    raw = render_docx(_ctx())
    assert raw[:2] == b"PK"
    paras = " ".join(p.text for p in Document(io.BytesIO(raw)).paragraphs if p.text)
    assert f"0 of {len(TECHNIQUES)} scored techniques cite at least one tool" in paras
    for clause in _METHODOLOGY_CLAUSES:
        assert clause in paras, f"methodology note lost from the DOCX: {clause!r}"
