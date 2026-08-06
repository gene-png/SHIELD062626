"""HTTP-level tests for the CSF deliverable workflow routes."""

from __future__ import annotations

import os
import uuid as _uuid
from collections.abc import Iterator
from pathlib import Path

import pytest
from alembic import command
from alembic.config import Config
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker

from app.storage.local import LocalFilesystemStorage


@pytest.fixture()
def app_client(tmp_path) -> Iterator[TestClient]:
    db_path = tmp_path / "shield-csfdeliv.db"
    url = f"sqlite:///{db_path}"
    os.environ["DATABASE_URL"] = url
    api_root = Path(__file__).resolve().parents[2]
    cfg = Config(str(api_root / "alembic.ini"))
    cfg.set_main_option("script_location", str(api_root / "alembic"))
    cfg.set_main_option("sqlalchemy.url", url)
    command.upgrade(cfg, "head")

    engine = create_engine(url, future=True)
    TestSession = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    storage = LocalFilesystemStorage(tmp_path / "storage")

    from app.db.session import get_db
    from app.main import create_app
    from app.routes.artifacts import _storage_dep

    def override_get_db() -> Iterator[Session]:
        db = TestSession()
        try:
            yield db
        finally:
            db.close()

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    app.dependency_overrides[_storage_dep] = lambda: storage
    # Multi-tenant (post-0013): admin/reviewer callers must name an active
    # tenant via X-Client-Id. Seed one tenant and bake the header into the
    # test client so single-tenant-style tests resolve to it; client-role
    # callers are pinned to their own client and ignore this header.
    from app.models.client import Client as _Client

    _seed = TestSession()
    _tenant = _Client(legal_name="Test Tenant")
    _seed.add(_tenant)
    _seed.flush()
    from app.models.client_domain import ClientDomain as _ClientDomain

    _seed.add(_ClientDomain(client_id=_tenant.id, domain="example.com"))
    _seed.commit()
    _cid = str(_tenant.id)
    _seed.close()

    with TestClient(app, headers={"X-Client-Id": _cid}) as c:
        yield c


def _register(c: TestClient, email: str) -> dict:
    r = c.post(
        "/auth/register",
        json={
            "email": email,
            "password": "correct horse battery staple!",
            "display_name": email.split("@")[0],
        },
    )
    assert r.status_code == 201, r.text
    return r.json()


def _seed_approved(c: TestClient, bearer: str) -> tuple[str, str]:
    """Open CSF service, create assessment, score everything tier 3, approve."""
    sr = c.post(
        "/csf/services",
        headers={"Authorization": f"Bearer {bearer}"},
        json={"kind": "nist_csf", "title": "Atlas - CSF"},
    )
    svc_id = sr.json()["id"]
    a = c.post(
        f"/csf/services/{svc_id}/assessments",
        headers={"Authorization": f"Bearer {bearer}"},
    )
    assess = a.json()
    # Patch every answer to tier 3 so the score is well-defined.
    for ans in assess["answers"]:
        c.patch(
            f"/csf/answers/{ans['id']}",
            headers={"Authorization": f"Bearer {bearer}"},
            json={"maturity_tier": 3},
        )
    c.post(
        f"/csf/assessments/{assess['id']}/approve",
        headers={"Authorization": f"Bearer {bearer}"},
    )
    return svc_id, assess["id"]


@pytest.mark.unit
def test_finalize_renders_pdf_and_xlsx(app_client) -> None:
    c = app_client
    admin = _register(c, "admin@example.com")
    bearer = admin["tokens"]["access_token"]
    svc_id, _ = _seed_approved(c, bearer)
    r = c.post(
        f"/csf/services/{svc_id}/deliverables/finalize",
        headers={"Authorization": f"Bearer {bearer}"},
    )
    assert r.status_code == 201, r.text
    body = r.json()
    assert body["version"] == 1
    assert body["finalized_at"] is not None
    assert body["pdf_artifact_id"] is not None
    assert body["xlsx_artifact_id"] is not None
    assert body["pdf_filename"].endswith(".pdf")
    assert "NIST_CSF_2_0_Assessment" in body["pdf_filename"]
    assert body["xlsx_filename"].endswith(".xlsx")
    assert "Overall maturity" in body["summary"]
    # Word deliverable (Work Order C4): present + a real .docx (zip starts "PK").
    assert body["docx_artifact_id"] is not None
    assert body["docx_filename"].endswith(".docx")
    docx = c.get(
        f"/artifacts/{body['docx_artifact_id']}/download",
        headers={"Authorization": f"Bearer {bearer}"},
    )
    assert docx.status_code == 200
    assert docx.content[:2] == b"PK"


@pytest.mark.unit
def test_finalize_requires_approved_assessment(app_client) -> None:
    c = app_client
    admin = _register(c, "admin@example.com")
    bearer = admin["tokens"]["access_token"]
    sr = c.post(
        "/csf/services",
        headers={"Authorization": f"Bearer {bearer}"},
        json={"kind": "nist_csf", "title": "Atlas"},
    )
    svc_id = sr.json()["id"]
    c.post(
        f"/csf/services/{svc_id}/assessments",
        headers={"Authorization": f"Bearer {bearer}"},
    )
    # No approve.
    r = c.post(
        f"/csf/services/{svc_id}/deliverables/finalize",
        headers={"Authorization": f"Bearer {bearer}"},
    )
    assert r.status_code == 409


@pytest.mark.unit
def test_finalize_404_for_non_csf_service(app_client) -> None:
    c = app_client
    admin = _register(c, "admin@example.com")
    bearer = admin["tokens"]["access_token"]
    td = c.post(
        "/tech-debt/services",
        headers={"Authorization": f"Bearer {bearer}"},
        json={"kind": "tech_debt", "title": "x"},
    )
    r = c.post(
        f"/csf/services/{td.json()['id']}/deliverables/finalize",
        headers={"Authorization": f"Bearer {bearer}"},
    )
    assert r.status_code == 404


@pytest.mark.unit
def test_latest_404_when_no_deliverable(app_client) -> None:
    c = app_client
    admin = _register(c, "admin@example.com")
    bearer = admin["tokens"]["access_token"]
    svc_id, _ = _seed_approved(c, bearer)
    r = c.get(
        f"/csf/services/{svc_id}/deliverables/latest",
        headers={"Authorization": f"Bearer {bearer}"},
    )
    assert r.status_code == 404


@pytest.mark.unit
def test_client_cannot_reach_csf_deliverable(app_client) -> None:
    """Work Order A1: clients never see or download deliverables in-app."""
    c = app_client
    admin = _register(c, "admin@example.com")
    bearer_admin = admin["tokens"]["access_token"]
    client = _register(c, "client@example.com")
    c.headers["X-Client-Id"] = client["user"]["client_id"]
    bearer_client = client["tokens"]["access_token"]
    svc_id, _ = _seed_approved(c, bearer_admin)
    fin = c.post(
        f"/csf/services/{svc_id}/deliverables/finalize",
        headers={"Authorization": f"Bearer {bearer_admin}"},
    )
    deliv = fin.json()
    # The latest-deliverable endpoint is admin-only (403 for a client).
    latest = c.get(
        f"/csf/services/{svc_id}/deliverables/latest",
        headers={"Authorization": f"Bearer {bearer_client}"},
    )
    assert latest.status_code == 403
    # The client cannot download the deliverable's artifacts (404, not theirs).
    pdf = c.get(
        f"/artifacts/{deliv['pdf_artifact_id']}/download",
        headers={"Authorization": f"Bearer {bearer_client}"},
    )
    assert pdf.status_code == 404


# ---------------------------------------------------------------------------
# S3: the finalize path loads the POA&M rows and passes them to build_context
# ---------------------------------------------------------------------------


def _seed_approved_with_gaps(c: TestClient, bearer: str) -> tuple[str, str, str]:
    """Like `_seed_approved` but every answer is tier 1, so the default T3
    target leaves a full gap list for the action plan to annotate."""
    sr = c.post(
        "/csf/services",
        headers={"Authorization": f"Bearer {bearer}"},
        json={"kind": "nist_csf", "title": "Atlas - CSF"},
    )
    svc_id = sr.json()["id"]
    a = c.post(
        f"/csf/services/{svc_id}/assessments",
        headers={"Authorization": f"Bearer {bearer}"},
    )
    assess = a.json()
    for ans in assess["answers"]:
        c.patch(
            f"/csf/answers/{ans['id']}",
            headers={"Authorization": f"Bearer {bearer}"},
            json={"maturity_tier": 1},
        )
    first_code = assess["answers"][0]["subcategory_code"]
    c.post(
        f"/csf/assessments/{assess['id']}/approve",
        headers={"Authorization": f"Bearer {bearer}"},
    )
    return svc_id, assess["id"], first_code


@pytest.mark.unit
def test_finalize_passes_the_gap_action_rows_into_build_context(app_client, monkeypatch) -> None:
    c = app_client
    admin = _register(c, "admin@example.com")
    bearer = admin["tokens"]["access_token"]
    svc_id, _assess_id, code = _seed_approved_with_gaps(c, bearer)
    saved = c.put(
        f"/csf/services/{svc_id}/gap-actions/{code}",
        headers={"Authorization": f"Bearer {bearer}"},
        json={
            "owner": "Priya Raman, CISO",
            "deadline": "2026-11-30",
            "priority_override": "P1",
        },
    )
    assert saved.status_code in (200, 201), saved.text

    from app.csf import exporters as csf_exporters
    from app.routes import csf as csf_routes

    seen: dict = {}

    def _spy(**kwargs):
        ctx = csf_exporters.build_context(**kwargs)
        # Snapshot inside the request: the ORM rows detach once it closes.
        seen["actions"] = {
            subcat: {
                "owner": row.owner,
                "deadline": row.deadline,
                "priority_override": row.priority_override,
            }
            for subcat, row in ctx.actions.items()
        }
        return ctx

    monkeypatch.setattr(csf_routes, "build_csf_context", _spy)

    r = c.post(
        f"/csf/services/{svc_id}/deliverables/finalize",
        headers={"Authorization": f"Bearer {bearer}"},
    )
    assert r.status_code == 201, r.text
    assert seen["actions"] == {
        code: {
            "owner": "Priya Raman, CISO",
            "deadline": "2026-11-30",
            "priority_override": "P1",
        }
    }


@pytest.mark.unit
def test_finalize_with_zero_gap_actions_still_renders(app_client) -> None:
    """C0: an assessment that never used the POA&M screen renders unchanged —
    the Gap Plan carries the new columns as empty cells and the action plan
    states the computed zero-owner line."""
    import io

    from openpyxl import load_workbook

    c = app_client
    admin = _register(c, "admin@example.com")
    bearer = admin["tokens"]["access_token"]
    svc_id, _assess_id, _code = _seed_approved_with_gaps(c, bearer)
    r = c.post(
        f"/csf/services/{svc_id}/deliverables/finalize",
        headers={"Authorization": f"Bearer {bearer}"},
    )
    assert r.status_code == 201, r.text
    body = r.json()

    xlsx = c.get(
        f"/artifacts/{body['xlsx_artifact_id']}/download",
        headers={"Authorization": f"Bearer {bearer}"},
    )
    ws = load_workbook(io.BytesIO(xlsx.content))["Gap Plan"]
    header = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    assert header[8:14] == [
        "Characterization",
        "Owner",
        "Deadline",
        "Resources",
        "Success criteria",
        "POA&M ref",
    ]
    # Every POA&M cell on the first gap row is blank (openpyxl round-trips the
    # written empty string back as an empty cell), not a placeholder or a crash.
    assert [ws.cell(row=2, column=col).value for col in range(9, 15)] == [None] * 6
    # Priority falls back to the code-computed score, formatted to 2dp.
    assert ws.cell(row=2, column=8).value == "2.40"

    pdf = c.get(
        f"/artifacts/{body['pdf_artifact_id']}/download",
        headers={"Authorization": f"Bearer {bearer}"},
    )
    from pypdf import PdfReader

    text = " ".join(
        "".join(p.extract_text() for p in PdfReader(io.BytesIO(pdf.content)).pages).split()
    )
    assert (
        "Start with the 20 subcategory gap(s) sitting 2 tier(s) below target "
        "T3 (Repeatable). They carry the largest lift." in text
    )
    assert "0 of 20 gap(s) in this action plan name an owner; assign the remaining 20." in text


@pytest.mark.unit
def test_finalize_404_for_unknown_service(app_client) -> None:
    c = app_client
    admin = _register(c, "admin@example.com")
    bearer = admin["tokens"]["access_token"]
    r = c.post(
        f"/csf/services/{_uuid.uuid4()}/deliverables/finalize",
        headers={"Authorization": f"Bearer {bearer}"},
    )
    assert r.status_code == 404
