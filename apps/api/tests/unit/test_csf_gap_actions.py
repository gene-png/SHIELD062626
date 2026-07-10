"""CSF POA&M / gap action plan (Sprint 5 T5, spec step 10).

Covers: default priority from the code-computed roll-up (gap_priority) with a
stored override winning; CRUD contract (admin-only, D-016 typed errors, autosave
upsert + reload survival); a fresh assessment parses with zero action rows (C0);
and the playbook XLSX gains an Action Plan sheet with the annotation columns.
"""

from __future__ import annotations

import io
import os
from collections.abc import Iterator
from pathlib import Path
from types import SimpleNamespace

import pytest
from alembic import command
from alembic.config import Config
from app.csf import playbook_export
from app.csf.catalog import SUBCATEGORIES
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker


@pytest.fixture()
def app_client(tmp_path) -> Iterator[TestClient]:
    url = f"sqlite:///{tmp_path / 'shield-csfpoam.db'}"
    os.environ["DATABASE_URL"] = url
    api_root = Path(__file__).resolve().parents[2]
    cfg = Config(str(api_root / "alembic.ini"))
    cfg.set_main_option("script_location", str(api_root / "alembic"))
    cfg.set_main_option("sqlalchemy.url", url)
    command.upgrade(cfg, "head")
    engine = create_engine(url, future=True)
    TestSession = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)

    from app.db.session import get_db
    from app.main import create_app
    from app.models.client import Client as _Client
    from app.models.client_domain import ClientDomain as _ClientDomain

    def override_get_db() -> Iterator[Session]:
        db = TestSession()
        try:
            yield db
        finally:
            db.close()

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    _seed = TestSession()
    tenant = _Client(legal_name="Test Tenant")
    _seed.add(tenant)
    _seed.flush()
    _seed.add(_ClientDomain(client_id=tenant.id, domain="example.com"))
    _seed.commit()
    cid = str(tenant.id)
    with TestClient(app, headers={"X-Client-Id": cid}) as c:
        yield c


def _register(c: TestClient, email: str) -> dict:
    r = c.post(
        "/auth/register",
        json={
            "email": email,
            "password": "correct horse battery staple!",
            "display_name": email.split("@")[0],
        },
    )
    assert r.status_code == 201, r.text
    return r.json()


def _admin_headers(c: TestClient) -> dict:
    bearer = _register(c, "admin@example.com")["tokens"]["access_token"]
    return {"Authorization": f"Bearer {bearer}"}


def _seed_gap(c: TestClient, h: dict) -> tuple[str, str]:
    """Open a CSF service, seed all three tiers, and force SUBCATEGORIES[0] into
    a gap (target L4, but every tier scored low). Returns (service_id, code)."""
    svc_id = c.post("/csf/services", headers=h, json={"kind": "nist_csf", "title": "CSF"}).json()[
        "id"
    ]
    c.post(f"/csf/services/{svc_id}/assessments", headers=h)
    c.post(
        f"/csf/services/{svc_id}/profiles/seed",
        headers=h,
        json={"tiers": ["high", "moderate", "low"]},
    )
    code = SUBCATEGORIES[0].code
    for tier in ("high", "moderate", "low"):
        rows = c.get(f"/csf/services/{svc_id}/profile/{tier}", headers=h).json()["rows"]
        sid = next(r["id"] for r in rows if r["subcategory_code"] == code)
        c.patch(
            f"/csf/dimension-scores/{sid}",
            headers=h,
            json={"has_evidence": True, "target_level": 4},
        )
    return svc_id, code


@pytest.mark.unit
def test_default_priority_from_engine_and_override_wins(app_client) -> None:
    c = app_client
    h = _admin_headers(c)
    svc_id, code = _seed_gap(c, h)

    # The code-computed default equals the Enterprise roll-up priority.
    ent = c.get(f"/csf/services/{svc_id}/enterprise-profile", headers=h).json()
    expected_default = next(s for s in ent["subcategories"] if s["subcategory_code"] == code)[
        "priority"
    ]
    assert expected_default in {"P1", "P2", "P3"}

    listing = c.get(f"/csf/services/{svc_id}/gap-actions", headers=h)
    assert listing.status_code == 200, listing.text
    actions = {a["subcategory_code"]: a for a in listing.json()["actions"]}
    assert code in actions
    a0 = actions[code]
    assert a0["default_priority"] == expected_default
    assert a0["effective_priority"] == expected_default
    # No annotation yet.
    assert a0["owner"] is None and a0["characterization"] is None
    assert a0["priority_override"] is None

    # Autosave one field at a time (owner, then characterization) — upsert.
    up = c.put(
        f"/csf/services/{svc_id}/gap-actions/{code}",
        headers=h,
        json={"owner": "Alice"},
    )
    assert up.status_code == 200, up.text
    assert up.json()["owner"] == "Alice"
    assert up.json()["default_priority"] == expected_default
    assert up.json()["effective_priority"] == expected_default

    c.put(
        f"/csf/services/{svc_id}/gap-actions/{code}",
        headers=h,
        json={"characterization": "mitigate", "poam_ref": "POAM-42"},
    )

    # A priority override wins for the effective value; the engine default stays.
    ov = c.put(
        f"/csf/services/{svc_id}/gap-actions/{code}",
        headers=h,
        json={"priority_override": "P1"},
    )
    assert ov.json()["priority_override"] == "P1"
    assert ov.json()["effective_priority"] == "P1"
    assert ov.json()["default_priority"] == expected_default

    # Reload survives (persisted, merged).
    reloaded = {
        a["subcategory_code"]: a
        for a in c.get(f"/csf/services/{svc_id}/gap-actions", headers=h).json()["actions"]
    }[code]
    assert reloaded["owner"] == "Alice"
    assert reloaded["characterization"] == "mitigate"
    assert reloaded["poam_ref"] == "POAM-42"
    assert reloaded["effective_priority"] == "P1"


@pytest.mark.unit
def test_typed_errors_on_bad_enum_and_unknown_subcategory(app_client) -> None:
    c = app_client
    h = _admin_headers(c)
    svc_id, code = _seed_gap(c, h)

    bad_ch = c.put(
        f"/csf/services/{svc_id}/gap-actions/{code}",
        headers=h,
        json={"characterization": "delegate"},
    )
    assert bad_ch.status_code == 422
    assert bad_ch.json()["error"]["reason"] == "invalid_characterization"

    bad_pri = c.put(
        f"/csf/services/{svc_id}/gap-actions/{code}",
        headers=h,
        json={"priority_override": "P9"},
    )
    assert bad_pri.status_code == 422
    assert bad_pri.json()["error"]["reason"] == "invalid_priority"

    unknown = c.put(
        f"/csf/services/{svc_id}/gap-actions/ZZ.ZZ-99",
        headers=h,
        json={"owner": "Bob"},
    )
    assert unknown.status_code == 404
    assert unknown.json()["error"]["reason"] == "unknown_subcategory"


@pytest.mark.unit
def test_gap_actions_admin_only(app_client) -> None:
    c = app_client
    h = _admin_headers(c)
    svc_id, code = _seed_gap(c, h)

    client = _register(c, "client@example.com")
    ch = {
        "Authorization": f"Bearer {client['tokens']['access_token']}",
        "X-Client-Id": client["user"]["client_id"],
    }
    assert c.get(f"/csf/services/{svc_id}/gap-actions", headers=ch).status_code == 403
    assert (
        c.put(
            f"/csf/services/{svc_id}/gap-actions/{code}",
            headers=ch,
            json={"owner": "X"},
        ).status_code
        == 403
    )


@pytest.mark.unit
def test_fresh_assessment_parses_with_zero_actions(app_client) -> None:
    # C0: the new table exists at head; a fresh assessment simply has no rows.
    c = app_client
    h = _admin_headers(c)
    svc_id, _code = _seed_gap(c, h)
    listing = c.get(f"/csf/services/{svc_id}/gap-actions", headers=h)
    assert listing.status_code == 200
    for a in listing.json()["actions"]:
        assert a["owner"] is None
        assert a["characterization"] is None
        assert a["priority_override"] is None


@pytest.mark.unit
def test_xlsx_gains_action_plan_sheet() -> None:
    rows = [
        SimpleNamespace(
            subcategory_code="GV.OC-01",
            name="Mission understood",
            function="GV",
            tier_levels={"high": 1},
            enterprise_level=1,
            rollup_rule=3,
            target_level=4,
            gap=True,
            priority="P2",
        ),
        SimpleNamespace(
            subcategory_code="GV.OC-02",
            name="No gap here",
            function="GV",
            tier_levels={"high": 5},
            enterprise_level=5,
            rollup_rule=1,
            target_level=4,
            gap=False,
            priority=None,
        ),
    ]
    gap_actions = {
        "GV.OC-01": SimpleNamespace(
            characterization="mitigate",
            priority_override="P1",
            owner="Alice",
            deadline="2026-09-30",
            resources="Two engineers",
            success_criteria="Policy signed",
            poam_ref="POAM-42",
        )
    }
    raw = playbook_export.render_xlsx(
        client_name="Acme",
        version=1,
        enterprise_rows=rows,
        tier_profiles={"high": []},
        gap_actions=gap_actions,
    )
    wb = load_workbook(io.BytesIO(raw))
    assert "Action Plan" in wb.sheetnames
    ws = wb["Action Plan"]
    header = [c.value for c in ws[1]]
    for col in (
        "Characterization",
        "Owner",
        "Deadline",
        "Resources",
        "Success criteria",
        "POA&M ref",
    ):
        assert col in header
    # Only the gap row appears, with the override priority + annotation values.
    body = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
    assert len(body) == 1
    row0 = dict(zip(header, body[0], strict=True))
    assert row0["Subcategory"] == "GV.OC-01"
    assert row0["Priority"] == "P1"  # override wins over the default P2
    assert row0["Characterization"] == "mitigate"
    assert row0["Owner"] == "Alice"
    assert row0["POA&M ref"] == "POAM-42"
