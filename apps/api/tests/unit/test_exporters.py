"""Filename slugifier + XLSX + PDF render smokes."""

from __future__ import annotations

import io
import uuid
from datetime import date

import pytest

from app.models.capability import (
    CapabilityDisposition,
    CapabilityItem,
    CapabilityList,
    CapabilityListStatus,
)
from app.tech_debt.exporters import (
    build_context,
    portfolio_paragraph,
    render_docx,
    render_pdf,
    render_xlsx,
)
from app.tech_debt.filename import (
    SERVICE_SLUG_TECH_DEBT,
    deliverable_filename,
    mmddyy,
    slugify,
)

# ---------------------------------------------------------------------------
# Slugifier
# ---------------------------------------------------------------------------


@pytest.mark.unit
@pytest.mark.parametrize(
    "raw,expected",
    [
        ("Acme, Inc.", "Acme_Inc"),
        ("  Atlas  Defense   Solutions  ", "Atlas_Defense_Solutions"),
        ("MITRE ATT&CK Coverage", "MITRE_ATTCK_Coverage"),
        ("___leading", "leading"),
        ("trailing___", "trailing"),
        ("", "Unknown"),
        ("   ", "Unknown"),
        ("!!!", "Unknown"),
        ("Nexus Federal Solutions Inc.", "Nexus_Federal_Solutions_Inc"),
        ("KEEP_CASE_AS_ENTERED", "KEEP_CASE_AS_ENTERED"),
        ("CamelCaseName", "CamelCaseName"),
    ],
)
def test_slugify(raw: str, expected: str) -> None:
    assert slugify(raw) == expected


@pytest.mark.unit
def test_slugify_none_returns_unknown() -> None:
    assert slugify(None) == "Unknown"


@pytest.mark.unit
def test_mmddyy_format() -> None:
    assert mmddyy(date(2026, 5, 18)) == "051826"
    assert mmddyy(date(2026, 1, 3)) == "010326"


@pytest.mark.unit
def test_deliverable_filename_matches_spec_example() -> None:
    # Spec §15.5: "Nexus Federal Solutions Inc. + Tech Debt Review + 2026-05-18"
    # -> "Nexus_Federal_Solutions_Inc_Tech_Debt_Review051826.pdf"
    name = deliverable_filename(
        company="Nexus Federal Solutions Inc.",
        service_slug=SERVICE_SLUG_TECH_DEBT,
        extension="pdf",
        day=date(2026, 5, 18),
    )
    assert name == "Nexus_Federal_Solutions_Inc_Tech_Debt_Review051826.pdf"


@pytest.mark.unit
def test_deliverable_filename_v2_re_release() -> None:
    name = deliverable_filename(
        company="Atlas Defense Solutions",
        service_slug=SERVICE_SLUG_TECH_DEBT,
        extension="xlsx",
        day=date(2026, 5, 18),
        version=2,
    )
    assert name == "Atlas_Defense_Solutions_Tech_Debt_Review051826_v2.xlsx"


@pytest.mark.unit
def test_deliverable_filename_working_prefix() -> None:
    name = deliverable_filename(
        company="X",
        service_slug=SERVICE_SLUG_TECH_DEBT,
        extension="xlsx",
        day=date(2026, 5, 18),
        working=True,
    )
    assert name == "WORKING_X_Tech_Debt_Review051826.xlsx"


# ---------------------------------------------------------------------------
# Render smokes
# ---------------------------------------------------------------------------


def _item(**kwargs) -> CapabilityItem:
    defaults = {
        "id": uuid.uuid4(),
        "capability_list_id": uuid.uuid4(),
        "name": "Wiz",
        "vendor": "Wiz, Inc.",
        "category": "CNAPP",
        "function": "Cloud posture",
        "annual_cost_usd": 350_000,
        "license_count": 200,
        "notes": None,
        "confidence_pct": 92,
        "source_artifact_id": None,
        "disposition": None,
        "disposition_rationale": None,
        "consolidation_target_id": None,
    }
    defaults.update(kwargs)
    return CapabilityItem(**defaults)


@pytest.fixture()
def context_with_items():
    cap_list = CapabilityList(
        id=uuid.uuid4(),
        service_id=uuid.uuid4(),
        version=1,
        status=CapabilityListStatus.APPROVED,
    )
    items = [
        _item(name="Wiz", annual_cost_usd=350_000, disposition=CapabilityDisposition.KEEP),
        _item(
            name="Lacework",
            annual_cost_usd=120_000,
            disposition=CapabilityDisposition.CUT,
        ),
        _item(
            name="Splunk",
            category="SIEM",
            annual_cost_usd=480_000,
            disposition=CapabilityDisposition.CONSOLIDATE,
        ),
    ]
    return build_context(
        client_legal_name="Atlas Defense Solutions",
        service_title="Technical Debt Review",
        cap_list=cap_list,
        items=items,
    )


@pytest.mark.unit
def test_xlsx_render_produces_valid_workbook(context_with_items) -> None:
    from openpyxl import load_workbook

    raw = render_xlsx(context_with_items)
    assert isinstance(raw, bytes)
    assert len(raw) > 1024  # non-trivial

    wb = load_workbook(io.BytesIO(raw))
    ws = wb.active
    assert ws.title == "Capability List"
    # Header row.
    assert ws.cell(row=1, column=1).value == "Name"
    assert ws.cell(row=1, column=5).value == "Annual Cost (USD)"
    # Data rows.
    row_values = [ws.cell(row=2, column=c).value for c in range(1, 11)]
    assert row_values[0] == "Wiz"
    assert row_values[4] == 350_000
    # Summary row total cost = 350k + 120k + 480k = 950k.
    found_total = False
    for row in ws.iter_rows(min_row=4):
        if row[0].value == "Total annual cost":
            assert row[4].value == 950_000
            found_total = True
    assert found_total


@pytest.mark.unit
def test_pdf_render_produces_valid_pdf(context_with_items) -> None:
    raw = render_pdf(context_with_items)
    assert isinstance(raw, bytes)
    # PDF magic.
    assert raw.startswith(b"%PDF-")
    # Reasonable size for a single-page report with a 3-row table.
    assert len(raw) > 1500


def _pdf_text(raw: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(raw))
    return "".join(page.extract_text() for page in reader.pages)


@pytest.mark.unit
def test_pdf_render_carries_title_client_and_a_known_row(context_with_items) -> None:
    # Upgrades the %PDF- magic smoke to a real content assertion (SMOKE §10):
    # the service title, the client name, and one known capability row survive
    # into the rendered PDF. Distinctive substrings only — reportlab text
    # extraction is whitespace-mangled, so we never assert exact layout.
    text = _pdf_text(render_pdf(context_with_items))
    assert "Technical Debt Review" in text  # service title
    assert "Atlas Defense Solutions" in text  # client name
    assert "Wiz" in text  # a known capability row


@pytest.mark.unit
def test_context_estimated_savings_sums_cut_items(context_with_items) -> None:
    # Only the Lacework row is "cut" with cost=120k.
    assert context_with_items.estimated_savings == 120_000
    assert context_with_items.savings_cost_known is True
    assert context_with_items.total_cost == 950_000


@pytest.mark.unit
def test_context_marks_savings_unknown_when_cut_has_no_cost() -> None:
    cap_list = CapabilityList(
        id=uuid.uuid4(), service_id=uuid.uuid4(), version=1, status=CapabilityListStatus.DRAFT
    )
    items = [
        _item(annual_cost_usd=None, disposition=CapabilityDisposition.CUT),
        _item(annual_cost_usd=50_000, disposition=CapabilityDisposition.CUT),
    ]
    ctx = build_context(client_legal_name=None, service_title="X", cap_list=cap_list, items=items)
    assert ctx.estimated_savings == 50_000
    assert ctx.savings_cost_known is False
    assert ctx.client_legal_name == "Client"  # fallback when None


# ---------------------------------------------------------------------------
# Portfolio narrative (Sprint 10 S5) - computed sentences only, and it must
# degrade honestly when the data is thin.
# ---------------------------------------------------------------------------


def _ctx(items, **kwargs):
    cap_list = CapabilityList(
        id=uuid.uuid4(),
        service_id=uuid.uuid4(),
        version=1,
        status=CapabilityListStatus.APPROVED,
    )
    return build_context(
        client_legal_name=kwargs.get("client", "Atlas Defense Solutions"),
        service_title="Technical Debt Review",
        cap_list=cap_list,
        items=items,
    )


@pytest.mark.unit
def test_portfolio_paragraph_counts_dispositions_and_names_cost_drivers(
    context_with_items,
) -> None:
    """Every number in the paragraph is computed from the rows, none is canned."""
    sentences = portfolio_paragraph(context_with_items)
    joined = " ".join(sentences)
    # 3 items: Wiz keep 350k, Lacework cut 120k, Splunk consolidate 480k.
    assert "3 capabilities" in joined
    assert "1 Keep" in joined and "1 Consolidate" in joined and "1 Cut" in joined
    # Top cost drivers, largest first, with their own costs.
    assert "Splunk ($480,000)" in joined
    assert joined.index("Splunk ($480,000)") < joined.index("Wiz ($350,000)")
    assert "$950,000" in joined  # recorded total
    # Savings framing: the cut row's cost, and its share of the recorded total.
    assert "$120,000" in joined
    assert "13% of the recorded total" in joined
    assert "Undecided" not in joined  # every row has a disposition


@pytest.mark.unit
def test_portfolio_paragraph_on_empty_capability_list_claims_nothing() -> None:
    """An empty list must say it is empty, never imply a reviewed portfolio."""
    sentences = portfolio_paragraph(_ctx([]))
    joined = " ".join(sentences)
    assert joined == (
        "No capabilities are recorded on this list, so there is no portfolio to "
        "summarize, no cost drivers to rank, and no savings to estimate."
    )
    # No dollar figure of any kind, so nothing can read as a finding.
    assert "$" not in joined


@pytest.mark.unit
def test_portfolio_paragraph_without_costs_reports_no_cost_drivers_or_savings() -> None:
    """Rows with no cost: no total, no drivers, no savings - stated, not implied."""
    items = [
        _item(name="Wiz", annual_cost_usd=None, disposition=CapabilityDisposition.KEEP),
        _item(name="Lacework", annual_cost_usd=None, disposition=CapabilityDisposition.CUT),
    ]
    joined = " ".join(portfolio_paragraph(_ctx(items)))
    assert "2 capabilities" in joined
    assert "No annual cost is recorded on any row" in joined
    assert "no cost drivers can be ranked and no savings can be estimated" in joined
    # Never a dollar figure, and never a savings claim, on absent cost data.
    assert "$" not in joined
    assert "lower bound" not in joined


@pytest.mark.unit
def test_portfolio_paragraph_without_dispositions_reports_no_split() -> None:
    """No disposition set anywhere: say the review is undecided, claim no savings."""
    items = [
        _item(name="Wiz", annual_cost_usd=350_000, disposition=None),
        _item(name="Splunk", annual_cost_usd=480_000, disposition=None),
    ]
    joined = " ".join(portfolio_paragraph(_ctx(items)))
    assert "None of the 2 capabilities carries a disposition yet" in joined
    assert "Keep" not in joined and "Consolidate" not in joined
    # Cost drivers still rank (that is recorded data)...
    assert "Splunk ($480,000)" in joined
    # ...but nothing is claimed as savings.
    assert "No row is marked Cut, so no annual savings are claimed." in joined


@pytest.mark.unit
def test_portfolio_paragraph_keeps_the_lower_bound_caveat_on_uncosted_cut() -> None:
    items = [
        _item(name="Wiz", annual_cost_usd=350_000, disposition=CapabilityDisposition.KEEP),
        _item(name="Lacework", annual_cost_usd=120_000, disposition=CapabilityDisposition.CUT),
        _item(name="Ghost", annual_cost_usd=None, disposition=CapabilityDisposition.CUT),
    ]
    joined = " ".join(portfolio_paragraph(_ctx(items)))
    assert "at least $120,000" in joined
    assert "1 Cut row carries no cost, so the figure is a lower bound" in joined
    # And the recorded-total sentence flags the missing cost rather than hiding it.
    assert "1 of the 3 rows carries no cost" in joined


@pytest.mark.unit
def test_portfolio_paragraph_reaches_pdf_docx_and_xlsx(context_with_items) -> None:
    """The criterion names all three surfaces, so assert all three carry it."""
    from openpyxl import load_workbook

    from app.docx_export import DOCX_MIME  # import guard: python-docx present

    assert DOCX_MIME
    first = portfolio_paragraph(context_with_items)[0]
    assert "3 capabilities" in first

    pdf_text = _pdf_text(render_pdf(context_with_items)).replace("\n", " ")
    assert "Portfolio" in pdf_text
    assert "1 Keep" in pdf_text

    ws = load_workbook(io.BytesIO(render_xlsx(context_with_items))).active
    xlsx_text = " ".join(
        str(c.value) for row in ws.iter_rows() for c in row if isinstance(c.value, str)
    )
    assert "Portfolio summary" in xlsx_text
    assert "1 Keep" in xlsx_text

    docx_text = _docx_text(render_docx(context_with_items))
    assert "Portfolio summary" in docx_text
    assert "1 Keep" in docx_text


def _docx_text(raw: bytes) -> str:
    import io as _io
    import zipfile

    with zipfile.ZipFile(_io.BytesIO(raw)) as z:
        return z.read("word/document.xml").decode("utf-8")


@pytest.mark.unit
def test_pdf_escapes_ampersand_in_portfolio_paragraph() -> None:
    """Capability names reach reportlab markup through the paragraph too."""
    items = [
        _item(name="Rook&Pawn Scanner", annual_cost_usd=10_000, disposition=None),
    ]
    text = _pdf_text(render_pdf(_ctx(items)))
    assert "Rook&Pawn Scanner" in text


@pytest.mark.unit
def test_pdf_escapes_ampersand_in_header() -> None:
    # The header lines feed reportlab Paragraph markup: a bare "&" in the
    # client name must be escaped so it renders literally instead of
    # re-emitting as an unknown entity with a synthesized semicolon.
    cap_list = CapabilityList(
        id=uuid.uuid4(),
        service_id=uuid.uuid4(),
        version=1,
        status=CapabilityListStatus.APPROVED,
    )
    ctx = build_context(
        client_legal_name="Rook&Pawn Security",
        service_title="Technical Debt Review",
        cap_list=cap_list,
        items=[_item(name="Wiz")],
    )
    text = _pdf_text(render_pdf(ctx))
    assert "Rook&Pawn Security" in text


# ---------------------------------------------------------------------------
# S9: PDF acceptance contract — section order plus one score/evidence/action link
# ---------------------------------------------------------------------------


def _pdf_norm(raw: bytes) -> str:
    """Extracted PDF text with reportlab's line wraps collapsed to spaces, so a
    sentence that happens to break across two lines still matches."""
    return " ".join(_pdf_text(raw).split())


def _assert_section_sequence(text: str, sequence: list[str]) -> None:
    """Assert every string appears in the rendered text, in exactly this order.

    Each needle is searched for only AFTER the previous match, so this is a true
    subsequence check rather than a set of `in` checks. A section that renders
    but lands in the wrong place still reads as a defensible report to `in`, and
    that is the failure this is here to catch. Reports the offending pair rather
    than a bare False.
    """
    last = -1
    last_needle = "(start of document)"
    for needle in sequence:
        found = text.find(needle, last + 1)
        assert found != -1, (
            f"{needle!r} does not appear in the rendered PDF after "
            f"{last_needle!r} (index {last})"
        )
        last, last_needle = found, needle


@pytest.mark.unit
def test_pdf_acceptance_contract_orders_sections_and_links_savings_to_its_row(
    context_with_items,
) -> None:
    """The Technical Debt PDF's acceptance contract, over real rendered bytes.

    Section order: Summary -> Portfolio summary -> Capability list.

    Representative linkage, all on the SAME $120,000: the headline savings
    figure (the score), the narrative sentence saying where it comes from (its
    evidence), and the capability row carrying the Cut disposition that produces
    it (its action). A reader has to be able to walk from the number to the row
    without leaving the document.
    """
    text = _pdf_norm(render_pdf(context_with_items))
    sentences = portfolio_paragraph(context_with_items)

    _assert_section_sequence(
        text,
        [
            # Section 1, and the score it states.
            "Summary",
            "Estimated annual savings: $120,000",
            # Section 2, and the evidence for that exact figure.
            "Portfolio summary",
            sentences[2],
            # Section 3, and the row whose disposition produces it.
            "Capability list",
            "Lacework Wiz, Inc. CNAPP $120,000 Cut",
        ],
    )
    # The evidence sentence is the savings sentence, not one of its neighbours —
    # a reordered paragraph would still satisfy the index walk above.
    assert "removes $120,000 of annual spend" in sentences[2]
