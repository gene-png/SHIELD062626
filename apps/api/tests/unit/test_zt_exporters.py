"""ZT PDF + XLSX exporter smokes for both frameworks."""

from __future__ import annotations

import io
import uuid

import pytest

from app.models.zt_assessment import (
    ZtAnswer,
    ZtAssessment,
    ZtAssessmentStatus,
    ZtFramework,
)
from app.zt.catalog import capabilities
from app.zt.exporters import build_context, render_pdf, render_xlsx
from app.zt.maturity import ZtFrameworkCode
from app.zt.scoring import analyze_gaps, compute


def _build_inputs(
    framework: ZtFrameworkCode, *, stage: int | None = 3
) -> tuple[ZtAssessment, list[ZtAnswer]]:
    db_framework = (
        ZtFramework.CISA_ZTMM_2_0
        if framework == ZtFrameworkCode.CISA_ZTMM_2_0
        else ZtFramework.DOD_ZTRA
    )
    a = ZtAssessment(
        id=uuid.uuid4(),
        service_id=uuid.uuid4(),
        framework=db_framework,
        version=1,
        status=ZtAssessmentStatus.APPROVED,
    )
    answers: list[ZtAnswer] = []
    for cap in capabilities(framework):
        answers.append(
            ZtAnswer(
                id=uuid.uuid4(),
                assessment_id=a.id,
                capability_code=cap.code,
                maturity_stage=stage,
            )
        )
    return a, answers


def _ctx(
    framework: ZtFrameworkCode,
    stage: int | None = 3,
    *,
    target: int = 4,
    narratives: dict | None = None,
    executive_summary: str | None = None,
    roadmap_summary: str | None = None,
    evidence_names: dict | None = None,
    scored_codes: set[str] | None = None,
):
    a, answers = _build_inputs(framework, stage=stage)
    if scored_codes is not None:
        for ans in answers:
            if ans.capability_code not in scored_codes:
                ans.maturity_stage = None
    a.pillar_narratives = narratives
    a.executive_summary = executive_summary
    a.roadmap_summary = roadmap_summary
    stage_map = {ans.capability_code: ans.maturity_stage for ans in answers}
    score = compute(framework, stage_map)
    gap = analyze_gaps(framework, stage_map, target_stage=target)
    return build_context(
        client_legal_name="Atlas Defense Solutions",
        service_title="Zero Trust Assessment",
        framework=framework,
        assessment=a,
        answers=answers,
        score=score,
        gap=gap,
        evidence_names=evidence_names,
    )


@pytest.mark.unit
def test_cisa_xlsx_has_three_sheets() -> None:
    from openpyxl import load_workbook

    raw = render_xlsx(_ctx(ZtFrameworkCode.CISA_ZTMM_2_0))
    assert raw[:2] == b"PK"
    wb = load_workbook(io.BytesIO(raw))
    assert set(wb.sheetnames) == {"Score Summary", "Answers", "Gap Plan"}


@pytest.mark.unit
def test_dod_xlsx_renders() -> None:
    from openpyxl import load_workbook

    raw = render_xlsx(_ctx(ZtFrameworkCode.DOD_ZTRA))
    wb = load_workbook(io.BytesIO(raw))
    assert "Score Summary" in wb.sheetnames
    ws = wb["Score Summary"]
    # Spot-check the framework cell.
    rows = [(ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value) for r in range(1, 8)]
    fw = dict(rows).get("Framework")
    assert fw == "DoD ZT Reference Architecture"


@pytest.mark.unit
def test_cisa_xlsx_answers_sheet_row_count() -> None:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(render_xlsx(_ctx(ZtFrameworkCode.CISA_ZTMM_2_0))))
    ws = wb["Answers"]
    # 37 capabilities + 1 header.
    assert ws.max_row == 38


@pytest.mark.unit
def test_dod_xlsx_answers_sheet_row_count() -> None:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(render_xlsx(_ctx(ZtFrameworkCode.DOD_ZTRA))))
    ws = wb["Answers"]
    # 50 capabilities + 1 header.
    assert ws.max_row == 51


@pytest.mark.unit
def test_cisa_pdf_renders() -> None:
    raw = render_pdf(_ctx(ZtFrameworkCode.CISA_ZTMM_2_0))
    assert raw.startswith(b"%PDF-")
    assert len(raw) > 2000


def _pdf_text(raw: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(raw))
    return "".join(page.extract_text() for page in reader.pages)


@pytest.mark.unit
def test_cisa_pdf_carries_title_client_framework_and_a_known_pillar() -> None:
    # SMOKE §10: upgrade from %PDF- magic to real content — title, client,
    # framework label, and one known per-pillar rollup row.
    text = _pdf_text(render_pdf(_ctx(ZtFrameworkCode.CISA_ZTMM_2_0)))
    assert "Zero Trust Assessment" in text  # service title
    assert "Atlas Defense Solutions" in text  # client name
    assert "CISA ZTMM 2.0" in text  # framework label
    assert "Identity" in text  # a known CISA pillar rollup row


@pytest.mark.unit
def test_dod_pdf_renders() -> None:
    raw = render_pdf(_ctx(ZtFrameworkCode.DOD_ZTRA))
    assert raw.startswith(b"%PDF-")


@pytest.mark.unit
def test_pdf_handles_empty_gap_list() -> None:
    # Score everyone Optimal (stage 4) and target 3 -> zero gaps.
    ctx = _ctx(ZtFrameworkCode.CISA_ZTMM_2_0, stage=4, target=3)
    raw = render_pdf(ctx)
    assert raw.startswith(b"%PDF-")


@pytest.mark.unit
def test_xlsx_handles_empty_gap_list_with_placeholder() -> None:
    from openpyxl import load_workbook

    ctx = _ctx(ZtFrameworkCode.CISA_ZTMM_2_0, stage=4, target=3)
    wb = load_workbook(io.BytesIO(render_xlsx(ctx)))
    ws = wb["Gap Plan"]
    assert ws.max_row == 2
    assert ws.cell(row=2, column=3).value == "No gaps at target stage"


@pytest.mark.unit
def test_build_context_falls_back_when_client_none() -> None:
    a, answers = _build_inputs(ZtFrameworkCode.CISA_ZTMM_2_0)
    stage_map = {ans.capability_code: ans.maturity_stage for ans in answers}
    score = compute(ZtFrameworkCode.CISA_ZTMM_2_0, stage_map)
    gap = analyze_gaps(ZtFrameworkCode.CISA_ZTMM_2_0, stage_map)
    ctx = build_context(
        client_legal_name=None,
        service_title="x",
        framework=ZtFrameworkCode.CISA_ZTMM_2_0,
        assessment=a,
        answers=answers,
        score=score,
        gap=gap,
    )
    assert ctx.client_legal_name == "Client"


# ---------------------------------------------------------------------------
# S4: roadmap, narratives, framework-aware heatmap, evidence
# ---------------------------------------------------------------------------


@pytest.mark.unit
def test_pdf_roadmap_section_carries_month_capability_pillar_and_ramp() -> None:
    """The Roadmap section's contract: a month, the capability, its pillar,
    and the current-to-target ramp, all computed from build_roadmap(gap.gaps)."""
    from app.zt.exporters import ROADMAP_HEADING, roadmap_rows

    # Everyone at stage 1 against a target of 4 -> every capability is a gap.
    ctx = _ctx(ZtFrameworkCode.CISA_ZTMM_2_0, stage=1, target=4)
    rows = roadmap_rows(ctx)
    assert rows, "a fully-gapped assessment must produce a roadmap"
    first = rows[0]

    text = _pdf_text(render_pdf(ctx))
    assert ROADMAP_HEADING in text
    assert f"M{first.month}" in text
    assert first.code in text  # the capability
    assert first.pillar_code in text  # its pillar
    assert f"S{first.current_stage} → S{first.target_stage}" in text


@pytest.mark.unit
def test_pdf_renders_both_narrative_sections_when_persisted() -> None:
    from app.zt.exporters import (
        CONSULTANT_HEADING,
        NARRATIVE_HEADING,
        NARRATIVE_METHODOLOGY_NOTE,
    )

    ctx = _ctx(
        ZtFrameworkCode.CISA_ZTMM_2_0,
        stage=2,
        narratives={"ID": "Identity is federated but not risk-adaptive."},
        executive_summary="Atlas is mid-transition on four of seven pillars.",
        roadmap_summary="Sequence identity first, then network segmentation.",
    )
    text = _pdf_text(render_pdf(ctx))
    assert NARRATIVE_HEADING in text
    assert "Atlas is mid-transition on four of seven pillars." in text
    assert "Identity is federated but not risk-adaptive." in text
    assert CONSULTANT_HEADING in text
    assert "Sequence identity first, then network segmentation." in text
    # AI attribution lives in the methodology note, not in the section labels.
    assert NARRATIVE_METHODOLOGY_NOTE[0][:60] in text


@pytest.mark.unit
def test_pdf_omits_narrative_headers_entirely_when_null() -> None:
    from app.zt.exporters import CONSULTANT_HEADING, NARRATIVE_HEADING

    ctx = _ctx(ZtFrameworkCode.CISA_ZTMM_2_0, stage=2)
    assert ctx.assessment.executive_summary is None
    assert ctx.assessment.pillar_narratives is None
    assert ctx.assessment.roadmap_summary is None
    text = _pdf_text(render_pdf(ctx))
    # Absent means omitted, not an empty header.
    assert NARRATIVE_HEADING not in text
    assert CONSULTANT_HEADING not in text
    assert "How these narratives were produced" not in text


@pytest.mark.unit
def test_docx_omits_narrative_headers_entirely_when_null() -> None:
    import docx

    from app.zt.exporters import CONSULTANT_HEADING, NARRATIVE_HEADING, render_docx

    ctx = _ctx(ZtFrameworkCode.CISA_ZTMM_2_0, stage=2)
    doc = docx.Document(io.BytesIO(render_docx(ctx)))
    headings = {p.text for p in doc.paragraphs}
    assert NARRATIVE_HEADING not in headings
    assert CONSULTANT_HEADING not in headings


@pytest.mark.unit
def test_dod_stage_heatmap_uses_a_three_rung_ladder() -> None:
    """DoD is a 3-rung ladder, CISA a 4-rung one, so the same stage number is a
    different point on the ramp. A framework-blind heatmap would return the same
    hex for both."""
    from app.export_style import GRADED_RAMP_HEX
    from app.zt.exporters import stage_hex

    dod = [stage_hex(s, ZtFrameworkCode.DOD_ZTRA) for s in (1, 2, 3)]
    # 3 rungs spread across the 7-step ramp: bottom, exact middle, top.
    assert dod == [GRADED_RAMP_HEX[0], GRADED_RAMP_HEX[3], GRADED_RAMP_HEX[6]]
    # Stage 2 of 3 (mid) is NOT stage 2 of 4 (below mid) — the ladder matters.
    assert stage_hex(2, ZtFrameworkCode.DOD_ZTRA) != stage_hex(2, ZtFrameworkCode.CISA_ZTMM_2_0)
    # A stage that does not exist on the DoD ladder raises rather than clamping.
    with pytest.raises(ValueError):
        stage_hex(4, ZtFrameworkCode.DOD_ZTRA)


@pytest.mark.unit
def test_xlsx_answers_sheet_shades_stage_on_the_framework_ladder() -> None:
    from openpyxl import load_workbook

    from app.zt.exporters import stage_hex

    wb = load_workbook(io.BytesIO(render_xlsx(_ctx(ZtFrameworkCode.DOD_ZTRA, stage=3))))
    ws = wb["Answers"]
    expected = "FF" + stage_hex(3, ZtFrameworkCode.DOD_ZTRA).lstrip("#").upper()
    assert ws.cell(row=2, column=5).fill.start_color.rgb == expected


@pytest.mark.unit
def test_evidence_reference_resolves_a_filename() -> None:
    from openpyxl import load_workbook

    framework = ZtFrameworkCode.CISA_ZTMM_2_0
    a, answers = _build_inputs(framework, stage=3)
    art_id = uuid.uuid4()
    answers[0].evidence_artifact_id = art_id
    stage_map = {ans.capability_code: ans.maturity_stage for ans in answers}
    ctx = build_context(
        client_legal_name="Atlas",
        service_title="ZT",
        framework=framework,
        assessment=a,
        answers=answers,
        score=compute(framework, stage_map),
        gap=analyze_gaps(framework, stage_map),
        evidence_names={art_id: "identity-policy-v3.pdf"},
    )
    wb = load_workbook(io.BytesIO(render_xlsx(ctx)))
    ws = wb["Answers"]
    cited_row = next(
        r
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=1).value == answers[0].capability_code
    )
    assert ws.cell(row=cited_row, column=7).value == "identity-policy-v3.pdf"


@pytest.mark.unit
def test_evidence_reference_says_no_evidence_only_when_null() -> None:
    from openpyxl import load_workbook

    from app.zt.exporters import NO_EVIDENCE_REFERENCE

    ctx = _ctx(ZtFrameworkCode.CISA_ZTMM_2_0, stage=3)
    wb = load_workbook(io.BytesIO(render_xlsx(ctx)))
    ws = wb["Answers"]
    assert ws.cell(row=1, column=7).value == "Evidence"
    assert ws.cell(row=2, column=7).value == NO_EVIDENCE_REFERENCE


@pytest.mark.unit
def test_build_context_raises_on_an_unresolved_evidence_pointer() -> None:
    """A pointer the route did not resolve must never degrade into the NULL
    sentence — that would be a false statement about the engagement."""
    framework = ZtFrameworkCode.CISA_ZTMM_2_0
    a, answers = _build_inputs(framework, stage=3)
    answers[0].evidence_artifact_id = uuid.uuid4()
    stage_map = {ans.capability_code: ans.maturity_stage for ans in answers}
    with pytest.raises(ValueError, match="missing 1 cited artifact"):
        build_context(
            client_legal_name="Atlas",
            service_title="ZT",
            framework=framework,
            assessment=a,
            answers=answers,
            score=compute(framework, stage_map),
            gap=analyze_gaps(framework, stage_map),
            evidence_names={},
        )


# ---------------------------------------------------------------------------
# Empty / sparse input: nothing may reassure about data that is not there
# ---------------------------------------------------------------------------


def _flat_pdf_text(raw: bytes) -> str:
    """PDF text with reportlab's line wrapping collapsed, so a sentence that
    happens to break across two lines still matches."""
    return " ".join(_pdf_text(raw).split())


@pytest.mark.unit
def test_unscored_assessment_never_reads_as_a_clean_bill_of_health() -> None:
    ctx = _ctx(ZtFrameworkCode.CISA_ZTMM_2_0, stage=None)
    assert ctx.score.answered_capabilities == 0
    text = _flat_pdf_text(render_pdf(ctx))
    assert "absence of data, not an absence of gaps" in text
    assert "empty for want of input, not because the work is complete" in text
    # The reassuring wording the fully-scored case uses must be absent here.
    assert "across all 37 capabilities" not in text


@pytest.mark.unit
def test_sparsely_scored_assessment_states_its_own_coverage() -> None:
    framework = ZtFrameworkCode.CISA_ZTMM_2_0
    scored = {c.code for c in capabilities(framework)[:3]}
    ctx = _ctx(framework, stage=4, target=3, scored_codes=scored)
    assert ctx.score.answered_capabilities == 3
    text = _flat_pdf_text(render_pdf(ctx))
    assert "among the 3 of 37 capabilities scored" in text
    assert "34 are unscored" in text
    # The headline "Overall stage: Optimal" is true of three capabilities and
    # misleading about the engagement, so the exclusion is stated in words.
    assert "Overall stage: Optimal" in text
    assert "The other 34 are unscored and are excluded from every average" in text


@pytest.mark.unit
def test_fully_scored_assessment_carries_no_coverage_qualifier() -> None:
    from app.zt.exporters import coverage_qualifier

    full = _ctx(ZtFrameworkCode.CISA_ZTMM_2_0, stage=4, target=3)
    assert full.score.answered_capabilities == full.score.total_capabilities
    assert coverage_qualifier(full) is None
    # Nor when nothing is scored: "Unscored" is already the honest headline.
    assert coverage_qualifier(_ctx(ZtFrameworkCode.CISA_ZTMM_2_0, stage=None)) is None
    assert "excluded from every average" not in _flat_pdf_text(render_pdf(full))


@pytest.mark.unit
def test_xlsx_gap_placeholder_names_the_unscored_case() -> None:
    from openpyxl import load_workbook

    ctx = _ctx(ZtFrameworkCode.CISA_ZTMM_2_0, stage=None)
    wb = load_workbook(io.BytesIO(render_xlsx(ctx)))
    assert wb["Gap Plan"].cell(row=2, column=3).value == "No capability scored — gaps unknown"


@pytest.mark.unit
def test_pdf_escapes_ampersand_in_header() -> None:
    # The header lines feed reportlab Paragraph markup: a bare "&" in the
    # client name (or title) must be escaped so it renders literally instead
    # of re-emitting as an unknown entity with a synthesized semicolon.
    framework = ZtFrameworkCode.CISA_ZTMM_2_0
    a, answers = _build_inputs(framework, stage=3)
    stage_map = {ans.capability_code: ans.maturity_stage for ans in answers}
    score = compute(framework, stage_map)
    gap = analyze_gaps(framework, stage_map, target_stage=4)
    ctx = build_context(
        client_legal_name="Rook&Pawn Security",
        service_title="Zero Trust (CISA ZTMM)",
        framework=framework,
        assessment=a,
        answers=answers,
        score=score,
        gap=gap,
    )
    text = _pdf_text(render_pdf(ctx))
    assert "Rook&Pawn Security" in text


# ---------------------------------------------------------------------------
# S9: PDF acceptance contract — section order plus one score/gap/roadmap link
# ---------------------------------------------------------------------------


def _assert_section_sequence(text: str, sequence: list[str]) -> None:
    """Assert every string appears in the rendered text, in exactly this order.

    Each needle is searched for only AFTER the previous match, so this is a true
    subsequence check rather than a set of `in` checks. A section that renders
    but lands in the wrong place still reads as a defensible report to `in`, and
    that is the failure this is here to catch.
    """
    last = -1
    last_needle = "(start of document)"
    for needle in sequence:
        found = text.find(needle, last + 1)
        assert found != -1, (
            f"{needle!r} does not appear in the rendered PDF after "
            f"{last_needle!r} (index {last})"
        )
        last, last_needle = found, needle


@pytest.mark.unit
def test_pdf_acceptance_contract_orders_sections_and_links_the_stage_to_its_roadmap() -> None:
    """The Zero Trust PDF's acceptance contract, over real rendered bytes.

    Section order: Maturity summary -> Per-pillar rollup -> Top remediation gaps
    -> Remediation roadmap -> Assessment narrative -> Consultant summary -> How
    these narratives were produced.

    Representative linkage, all on the SAME capability: the average stage (the
    score), the gap row showing its ramp to target, and the roadmap row that
    schedules the month it gets closed (its action). The narrative attribution is
    the citation reference, and it has to come LAST — a reader who has already
    read the prose needs to be told, in the same document, that none of the
    numbers above came from it.
    """
    from app.zt.exporters import (
        CONSULTANT_HEADING,
        NARRATIVE_HEADING,
        ROADMAP_HEADING,
        roadmap_rows,
    )

    ctx = _ctx(
        ZtFrameworkCode.CISA_ZTMM_2_0,
        stage=1,
        target=4,
        narratives={"identity": "Identity is perimeter-shaped today."},
        executive_summary="Atlas is early on the CISA ladder.",
        roadmap_summary="Sequence identity first, then devices.",
    )
    text = _flat_pdf_text(render_pdf(ctx))
    gap = ctx.gap.gaps[0]
    first = roadmap_rows(ctx)[0]
    assert first.code == gap.code, "the roadmap must open on the top-priority gap"

    _assert_section_sequence(
        text,
        [
            # Section 1, and the score.
            "Maturity summary",
            "Overall stage: Traditional",
            "Average stage: 1.00",
            # Section 2.
            "Per-pillar rollup",
            # Section 3, and the gap's ramp on this capability.
            f"Top remediation gaps (target S{ctx.gap.target_stage})",
            f"{gap.code} {gap.pillar_code}",
            f"S{gap.current_stage} → S{gap.target_stage}",
            # Section 4, and the month that same capability is scheduled for.
            ROADMAP_HEADING,
            f"M{first.month} {first.code}",
            # Sections 5 and 6: the prose, rendered only because it is persisted.
            NARRATIVE_HEADING,
            "Atlas is early on the CISA ladder.",
            CONSULTANT_HEADING,
            "Sequence identity first, then devices.",
            # Section 7: the attribution, after the prose it qualifies.
            "How these narratives were produced",
            "No narrative in this report contributes to any of those numbers.",
        ],
    )
