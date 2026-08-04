"""Zero Trust deliverable renderers - PDF + XLSX from a ZT assessment.

Three-sheet XLSX (Score Summary / Answers / Gap Plan) + executive PDF
(overall stage + per-pillar table, then top-N gap table, the sequenced
roadmap, and the persisted narratives). Pure functions, no I/O.
Framework-aware: CISA labels render Traditional/Initial/..., DoD labels
render Not Started/Target/..., and every graded fill is
`graded_hex(stage, level_count(framework))` so a DoD 3 is the top of a
3-rung ladder and a CISA 4 the top of a 4-rung one.

S4 note on language. The narratives are drafted by the zt_score Run AI and
persisted on the assessment (migration 0034). They are prose only: every stage,
average, gap, priority and roadmap month in this document is computed in Python
from the recorded scores, and no narrative feeds a number. Two rules follow, and
both are load-bearing:

* A narrative section renders only when its field is non-NULL. An empty header
  would imply a consultant left a section blank.
* No sentence here may read as reassurance about data that was never entered. A
  zero-gap finding across four scored capabilities out of thirty-seven is a
  coverage statement, not a clean bill of health, and it is printed as one.
"""

from __future__ import annotations

import io
import logging
import uuid
from collections.abc import Iterable, Mapping
from dataclasses import dataclass, field
from typing import TYPE_CHECKING, Any

from app import export_style
from app.models.zt_assessment import ZtAnswer, ZtAssessment
from app.zt.catalog import capabilities, pillars
from app.zt.maturity import ZtFrameworkCode, level_count, stage_label
from app.zt.scoring import GapAnalysis, RoadmapItem, ScoreResult, build_roadmap

if TYPE_CHECKING:
    from reportlab.platypus import TableStyle

logger = logging.getLogger(__name__)
_LOG = "zt.exporters:"

# Printed when `evidence_artifact_id` is genuinely NULL. An unresolved pointer
# raises in `build_context`, so this sentence never covers a failed lookup.
NO_EVIDENCE_REFERENCE = "No evidence attached"

# The two persisted-narrative section labels. Deliberately not "Executive
# summary": the attribution for both lives in NARRATIVE_METHODOLOGY_NOTE.
NARRATIVE_HEADING = "Assessment narrative"
CONSULTANT_HEADING = "Consultant summary"

# AI attribution, the way the CSF playbook carries its METHODOLOGY block: one
# place, printed with the narratives and only with them.
NARRATIVE_METHODOLOGY_NOTE: tuple[str, ...] = (
    "The narrative sections in this report are drafted by Run AI from the "
    "maturity stages and consultant notes recorded against this engagement, and "
    "are carried into the report only from an assessment a consultant approved.",
    "Every stage, average, coverage percentage, gap, priority score and roadmap "
    "month is computed in code from the recorded stages. No narrative in this "
    "report contributes to any of those numbers.",
)


@dataclass(frozen=True)
class ZtDeliverableContext:
    client_legal_name: str
    service_title: str
    framework: ZtFrameworkCode
    assessment: ZtAssessment
    answers: list[ZtAnswer]
    score: ScoreResult
    gap: GapAnalysis
    # artifact id -> filename, resolved by the route's tenant-filtered join.
    # Every non-NULL `evidence_artifact_id` on `answers` must appear here.
    evidence_names: dict[uuid.UUID, str] = field(default_factory=dict)


def build_context(
    *,
    client_legal_name: str | None,
    service_title: str,
    framework: ZtFrameworkCode,
    assessment: ZtAssessment,
    answers: Iterable[ZtAnswer],
    score: ScoreResult,
    gap: GapAnalysis,
    evidence_names: Mapping[uuid.UUID, str] | None = None,
) -> ZtDeliverableContext:
    rows = list(answers)
    names = dict(evidence_names or {})
    unresolved = sorted(
        str(r.evidence_artifact_id)
        for r in rows
        if r.evidence_artifact_id is not None and r.evidence_artifact_id not in names
    )
    if unresolved:
        raise ValueError(
            f"{_LOG} evidence_names is missing {len(unresolved)} cited artifact id(s): "
            f"{unresolved[:5]}. Rendering '{NO_EVIDENCE_REFERENCE}' for an answer that "
            f"does cite evidence would misstate the record."
        )
    logger.debug("%s build_context answers=%d evidence_names=%d", _LOG, len(rows), len(names))
    return ZtDeliverableContext(
        client_legal_name=client_legal_name or "Client",
        service_title=service_title,
        framework=framework,
        assessment=assessment,
        answers=rows,
        score=score,
        gap=gap,
        evidence_names=names,
    )


def _fmt(value: float | None) -> str:
    if value is None:
        return "—"
    return f"{value:.2f}"


def _framework_label(framework: ZtFrameworkCode) -> str:
    return (
        "CISA ZTMM 2.0"
        if framework == ZtFrameworkCode.CISA_ZTMM_2_0
        else "DoD ZT Reference Architecture"
    )


# ---------------------------------------------------------------------------
# Evidence
# ---------------------------------------------------------------------------


def evidence_reference(answer: ZtAnswer | None, names: Mapping[uuid.UUID, str]) -> str:
    """The attached artifact's filename, or the NULL sentence. Never a lookup miss."""
    if answer is None or answer.evidence_artifact_id is None:
        return NO_EVIDENCE_REFERENCE
    try:
        return names[answer.evidence_artifact_id]
    except KeyError as exc:
        raise KeyError(
            f"{_LOG} capability {answer.capability_code} cites evidence artifact "
            f"{answer.evidence_artifact_id} which the route did not resolve"
        ) from exc


# ---------------------------------------------------------------------------
# Stage heatmap — framework-aware (CISA 4 rungs, DoD 3)
# ---------------------------------------------------------------------------


def _argb(hex_color: str) -> str:
    """openpyxl wants ARGB. Derived from a brand hex, never hand-copied."""
    return "FF" + hex_color.lstrip("#").upper()


def stage_hex(stage: int, framework: ZtFrameworkCode) -> str:
    """Fill for a maturity stage on the framework's own ladder.

    `graded_hex` raises rather than clamping, so a DoD row carrying a stage 4 —
    which does not exist on a 3-rung ladder — fails loudly instead of colouring
    itself the top of a scale it is not on.
    """
    return export_style.graded_hex(stage, level_count(framework))


def stage_fill(stage: int, framework: ZtFrameworkCode) -> Any:
    from openpyxl.styles import PatternFill

    argb = _argb(stage_hex(stage, framework))
    return PatternFill(start_color=argb, end_color=argb, fill_type="solid")


def stage_font(stage: int, framework: ZtFrameworkCode) -> Any:
    """AA-safe font to print on `stage_fill(stage, framework)`."""
    from openpyxl.styles import Font

    return Font(color=_argb(export_style.graded_ink_hex(stage, level_count(framework))))


# ---------------------------------------------------------------------------
# Honest coverage language (the S3 false-reassurance defect, avoided)
# ---------------------------------------------------------------------------


def no_gap_sentence(ctx: ZtDeliverableContext) -> str:
    """What an empty gap list means, qualified by how much was actually scored.

    Zero gaps across zero scored capabilities is an absence of data. Saying
    anything else would be the reassurance defect S3 shipped.
    """
    scored = ctx.score.answered_capabilities
    total = ctx.score.total_capabilities
    target = f"stage {ctx.gap.target_stage} ({ctx.gap.target_label})"
    if scored == 0:
        return (
            f"No capability in this assessment has been scored, so no gap to target "
            f"{target} can be identified. This is an absence of data, not an absence "
            f"of gaps."
        )
    if scored < total:
        return (
            f"No gap to target {target} among the {scored} of {total} capabilities "
            f"scored. The remaining {total - scored} are unscored and this statement "
            f"says nothing about them."
        )
    return f"No gaps to target {target} across all {total} capabilities."


def coverage_qualifier(ctx: ZtDeliverableContext) -> str | None:
    """What the headline stage does and does not cover. None when it covers all.

    "Overall stage: Optimal" printed against three scored capabilities out of
    thirty-seven is true of the three and misleading about the engagement. The
    coverage fraction alone did not stop a reader taking the label at face value,
    so the exclusion is stated in words.
    """
    scored = ctx.score.answered_capabilities
    total = ctx.score.total_capabilities
    if scored == 0 or scored >= total:
        return None
    return (
        f"The overall and per-pillar stages above are computed from the {scored} of "
        f"{total} capabilities that carry a stage. The other {total - scored} are "
        f"unscored and are excluded from every average on this page, so no stage here "
        f"describes them."
    )


def no_roadmap_sentence(ctx: ZtDeliverableContext) -> str:
    """What an empty roadmap means. Never 'the work is done'."""
    scored = ctx.score.answered_capabilities
    total = ctx.score.total_capabilities
    if scored == 0:
        return (
            "No capability has been scored, so there is nothing to sequence. This "
            "roadmap is empty for want of input, not because the work is complete."
        )
    if scored < total:
        return (
            f"Nothing to sequence: none of the {scored} of {total} capabilities scored "
            f"sits below its target stage. The {total - scored} unscored capabilities "
            f"are not represented here."
        )
    return f"Nothing to sequence: all {total} capabilities are at or above their target " f"stage."


ROADMAP_HEADING = "Remediation roadmap (12 months)"
ROADMAP_COLUMNS: tuple[str, ...] = (
    "Month",
    "Capability",
    "Name",
    "Pillar",
    "Current → Target",
)


def roadmap_rows(ctx: ZtDeliverableContext) -> tuple[RoadmapItem, ...]:
    """The sequenced remediation plan. Computed in code from the gap list."""
    rows = build_roadmap(ctx.gap.gaps)
    logger.debug("%s roadmap_rows gaps=%d items=%d", _LOG, len(ctx.gap.gaps), len(rows))
    return rows


def _roadmap_table_rows(roadmap: Iterable[RoadmapItem]) -> list[list[str]]:
    return [
        [
            f"M{it.month}",
            it.code,
            it.name,
            f"{it.pillar_code} · {it.pillar_name}",
            f"S{it.current_stage} → S{it.target_stage}",
        ]
        for it in roadmap
    ]


# ---------------------------------------------------------------------------
# Persisted narratives (migration 0034) — rendered only when present
# ---------------------------------------------------------------------------


def _pillar_narrative_lines(ctx: ZtDeliverableContext) -> list[str]:
    raw = ctx.assessment.pillar_narratives or {}
    names = {p.code: p.name for p in pillars(ctx.framework)}
    return [
        f"{code} — {names.get(code, code)}: {text}"
        for code, text in sorted(raw.items())
        if str(text).strip()
    ]


def narrative_lines(ctx: ZtDeliverableContext) -> list[str]:
    """Body of the "Assessment narrative" section, empty when nothing persisted."""
    lines: list[str] = []
    summary = (ctx.assessment.executive_summary or "").strip()
    if summary:
        lines.append(summary)
    lines.extend(_pillar_narrative_lines(ctx))
    return lines


def consultant_summary_lines(ctx: ZtDeliverableContext) -> list[str]:
    """Body of the "Consultant summary" section, empty when nothing persisted."""
    summary = (ctx.assessment.roadmap_summary or "").strip()
    return [summary] if summary else []


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def render_xlsx(ctx: ZtDeliverableContext) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    header_fill = export_style.xlsx_header_fill()
    bold = Font(bold=True)
    italic = Font(italic=True)

    # --- Score Summary ---
    ws = wb.create_sheet("Score Summary")
    ws.append(["Engagement", ctx.client_legal_name])
    ws.append(["Service", ctx.service_title])
    ws.append(["Framework", _framework_label(ctx.framework)])
    ws.append(["Assessment version", ctx.assessment.version])
    ws.append(["Overall stage", ctx.score.overall_stage_label])
    ws.append(["Average stage", _fmt(ctx.score.average_stage)])
    ws.append(["Coverage", f"{ctx.score.answered_capabilities}/{ctx.score.total_capabilities}"])
    for row in ws.iter_rows(min_row=1, max_row=7, min_col=1, max_col=1):
        for cell in row:
            cell.font = bold
    ws.append([])
    ws.append(["Pillar", "Name", "Answered", "Total", "Coverage %", "Average stage"])
    for col_idx in range(1, 7):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.font = bold
        cell.fill = header_fill
    for ps in ctx.score.by_pillar:
        ws.append(
            [
                ps.pillar_code,
                ps.pillar_name,
                ps.answered_count,
                ps.capability_count,
                ps.coverage_pct,
                _fmt(ps.average_stage),
            ]
        )
    for w, col in zip([10, 36, 12, 10, 14, 16], range(1, 7), strict=True):
        ws.column_dimensions[get_column_letter(col)].width = w

    # --- Answers ---
    ws2 = wb.create_sheet("Answers")
    headers = [
        "Capability",
        "Pillar",
        "Name",
        "Outcome",
        "Stage",
        "Stage label",
        "Evidence",
        "Notes",
    ]
    ws2.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    answers_by_code = {a.capability_code: a for a in ctx.answers}
    # Use catalog order so missing answers still render as blank rows.
    pillar_lookup = {p.code: p.name for p in pillars(ctx.framework)}
    for cap in capabilities(ctx.framework):
        ans = answers_by_code.get(cap.code)
        s = ans.maturity_stage if ans else None
        notes = ans.notes if ans else None
        ws2.append(
            [
                cap.code,
                f"{cap.pillar_code} · {pillar_lookup.get(cap.pillar_code, cap.pillar_code)}",
                cap.name,
                cap.outcome,
                s if s is not None else "",
                stage_label(s, ctx.framework) if s is not None else "Unscored",
                evidence_reference(ans, ctx.evidence_names),
                notes or "",
            ]
        )
        # Heatmap on the framework's own ladder. An unscored row stays unfilled:
        # white is not a stage, and shading it would invent one.
        if s is not None:
            for col in (5, 6):
                shaded = ws2.cell(row=ws2.max_row, column=col)
                shaded.fill = stage_fill(s, ctx.framework)
                shaded.font = stage_font(s, ctx.framework)
    for w, col in zip([18, 30, 36, 60, 8, 16, 34, 60], range(1, 9), strict=True):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # --- Gap Plan ---
    ws3 = wb.create_sheet("Gap Plan")
    headers3 = [
        "Capability",
        "Pillar",
        "Name",
        "Current stage",
        "Target stage",
        "Gap size",
        "Priority",
        "Notes",
    ]
    ws3.append(headers3)
    for col in range(1, len(headers3) + 1):
        cell = ws3.cell(row=1, column=col)
        cell.font = bold
        cell.fill = header_fill
    for g in ctx.gap.gaps:
        ws3.append(
            [
                g.code,
                g.pillar_code,
                g.name,
                g.current_stage,
                g.target_stage,
                g.gap_size,
                g.priority_score,
                g.notes or "",
            ]
        )
    if not ctx.gap.gaps:
        # An empty gap list means different things at different coverage levels,
        # so the placeholder says which one. The fully-scored wording is the
        # original and stays byte-identical.
        scored = ctx.score.answered_capabilities
        total = ctx.score.total_capabilities
        if scored == 0:
            placeholder = "No capability scored — gaps unknown"
        elif scored < total:
            placeholder = f"No gaps among the {scored} of {total} capabilities scored"
        else:
            placeholder = "No gaps at target stage"
        ws3.append(["—", "", placeholder, "", ctx.gap.target_stage, 0, 0, ""])
        ws3.cell(row=2, column=3).font = italic
    for w, col in zip([18, 10, 36, 14, 14, 12, 12, 50], range(1, 9), strict=True):
        ws3.column_dimensions[get_column_letter(col)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def render_docx(ctx: ZtDeliverableContext) -> bytes:
    """Word deliverable mirroring the PDF (Work Order C4)."""
    from app.docx_export import (
        add_heading,
        add_paragraphs,
        add_table,
        add_title,
        new_document,
        to_bytes,
    )

    doc = new_document(export_style.metadata_title(ctx.service_title, ctx.client_legal_name))
    add_title(
        doc,
        ctx.service_title,
        f"{ctx.client_legal_name} · {_framework_label(ctx.framework)}",
    )

    add_heading(doc, "Maturity summary")
    add_paragraphs(
        doc,
        [
            f"Overall stage: {ctx.score.overall_stage_label}",
            f"Average stage: {_fmt(ctx.score.average_stage)}",
            f"Coverage: {ctx.score.answered_capabilities}/"
            f"{ctx.score.total_capabilities} ({ctx.score.coverage_pct}%)",
        ],
    )
    qualifier = coverage_qualifier(ctx)
    if qualifier:
        add_paragraphs(doc, [qualifier])

    add_heading(doc, "Per-pillar rollup")
    add_table(
        doc,
        ["Pillar", "Name", "Average stage", "Coverage"],
        [
            [
                ps.pillar_code,
                ps.pillar_name,
                _fmt(ps.average_stage),
                f"{ps.answered_count}/{ps.capability_count} ({ps.coverage_pct}%)",
            ]
            for ps in ctx.score.by_pillar
        ],
    )

    add_heading(doc, f"Top remediation gaps (target S{ctx.gap.target_stage})")
    if not ctx.gap.gaps:
        add_paragraphs(doc, [no_gap_sentence(ctx)])
    else:
        add_table(
            doc,
            ["Code", "Pillar", "Capability", "Current → Target", "Priority"],
            [
                [
                    g.code,
                    g.pillar_code,
                    g.name,
                    f"S{g.current_stage} → S{g.target_stage}",
                    f"{g.priority_score:.2f}",
                ]
                for g in ctx.gap.gaps
            ],
        )

    add_heading(doc, ROADMAP_HEADING)
    roadmap = roadmap_rows(ctx)
    if not roadmap:
        add_paragraphs(doc, [no_roadmap_sentence(ctx)])
    else:
        add_table(doc, list(ROADMAP_COLUMNS), _roadmap_table_rows(roadmap))

    narrative = narrative_lines(ctx)
    consultant = consultant_summary_lines(ctx)
    if narrative:
        add_heading(doc, NARRATIVE_HEADING)
        add_paragraphs(doc, narrative)
    if consultant:
        add_heading(doc, CONSULTANT_HEADING)
        add_paragraphs(doc, consultant)
    if narrative or consultant:
        add_heading(doc, "How these narratives were produced")
        add_paragraphs(doc, list(NARRATIVE_METHODOLOGY_NOTE))

    return to_bytes(doc)


def render_pdf(ctx: ZtDeliverableContext) -> bytes:
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        Spacer,
        Table,
    )

    out = io.BytesIO()
    doc = export_style.new_pdf_doc(
        out,
        title=export_style.metadata_title(ctx.service_title, ctx.client_legal_name),
        side_margin_in=export_style.SERVICE_PAGE_MARGIN_IN,
    )
    styles = getSampleStyleSheet()
    h1 = styles["Title"]
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], spaceBefore=14, spaceAfter=6)
    body = styles["BodyText"]

    story: list = []
    story.append(
        Paragraph(export_style.escaped_title(ctx.service_title, ctx.client_legal_name), h1)
    )
    story.append(
        Paragraph(
            f"{export_style.escaped_line(ctx.client_legal_name)} · "
            f"{_framework_label(ctx.framework)}",
            body,
        )
    )
    story.append(Spacer(1, 0.2 * inch))

    story.append(Paragraph("Maturity summary", h2))
    story.append(
        Paragraph(
            f"Overall stage: <b>{ctx.score.overall_stage_label}</b> · "
            f"Average stage: <b>{_fmt(ctx.score.average_stage)}</b> · "
            f"Coverage: <b>{ctx.score.answered_capabilities}/"
            f"{ctx.score.total_capabilities}</b> "
            f"({ctx.score.coverage_pct}%)",
            body,
        )
    )
    qualifier = coverage_qualifier(ctx)
    if qualifier:
        story.append(Paragraph(export_style.escaped_line(qualifier), body))

    story.append(Paragraph("Per-pillar rollup", h2))
    fn_table_data: list[list] = [["Pillar", "Name", "Average stage", "Coverage"]]
    for ps in ctx.score.by_pillar:
        fn_table_data.append(
            [
                ps.pillar_code,
                ps.pillar_name,
                _fmt(ps.average_stage),
                f"{ps.answered_count}/{ps.capability_count} ({ps.coverage_pct}%)",
            ]
        )
    fn_table = Table(
        fn_table_data,
        colWidths=[0.8 * inch, 3.2 * inch, 1.2 * inch, 1.6 * inch],
        repeatRows=1,
    )
    fn_table.setStyle(_table_style())
    story.append(fn_table)

    story.append(PageBreak())

    story.append(Paragraph(f"Top remediation gaps (target S{ctx.gap.target_stage})", h2))
    if not ctx.gap.gaps:
        story.append(Paragraph(export_style.escaped_line(no_gap_sentence(ctx)), body))
    else:
        gap_table_data: list[list] = [
            ["Code", "Pillar", "Capability", "Current → Target", "Priority"]
        ]
        for g in ctx.gap.gaps:
            gap_table_data.append(
                [
                    g.code,
                    g.pillar_code,
                    g.name,
                    f"S{g.current_stage} → S{g.target_stage}",
                    f"{g.priority_score:.2f}",
                ]
            )
        gap_table = Table(
            gap_table_data,
            colWidths=[1.1 * inch, 0.8 * inch, 2.9 * inch, 1.4 * inch, 0.8 * inch],
            repeatRows=1,
        )
        gap_table.setStyle(_table_style())
        story.append(gap_table)

    story.append(Paragraph(ROADMAP_HEADING, h2))
    roadmap = roadmap_rows(ctx)
    if not roadmap:
        story.append(Paragraph(export_style.escaped_line(no_roadmap_sentence(ctx)), body))
    else:
        roadmap_data: list[list] = [list(ROADMAP_COLUMNS)]
        roadmap_data.extend(_roadmap_table_rows(roadmap))
        roadmap_table = Table(
            roadmap_data,
            colWidths=[0.7 * inch, 1.1 * inch, 2.6 * inch, 1.8 * inch, 1.2 * inch],
            repeatRows=1,
        )
        roadmap_table.setStyle(_table_style())
        story.append(roadmap_table)

    # Narrative sections render only when the assessment actually carries them.
    # An empty header would read as a section a consultant left blank.
    narrative = narrative_lines(ctx)
    consultant = consultant_summary_lines(ctx)
    if narrative:
        story.append(Paragraph(NARRATIVE_HEADING, h2))
        for line in narrative:
            story.append(Paragraph(export_style.escaped_line(line), body))
    if consultant:
        story.append(Paragraph(CONSULTANT_HEADING, h2))
        for line in consultant:
            story.append(Paragraph(export_style.escaped_line(line), body))
    if narrative or consultant:
        story.append(Paragraph("How these narratives were produced", h2))
        for line in NARRATIVE_METHODOLOGY_NOTE:
            story.append(Paragraph(export_style.escaped_line(line), body))

    doc.build(story)
    return out.getvalue()


def _table_style() -> TableStyle:
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle

    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(export_style.SURFACE_SUNKEN_HEX)),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(export_style.INK_HEX)),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(export_style.BORDER_HEX)),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ]
    )


__all__ = [
    "CONSULTANT_HEADING",
    "NARRATIVE_HEADING",
    "NARRATIVE_METHODOLOGY_NOTE",
    "NO_EVIDENCE_REFERENCE",
    "ROADMAP_COLUMNS",
    "ROADMAP_HEADING",
    "ZtDeliverableContext",
    "build_context",
    "consultant_summary_lines",
    "coverage_qualifier",
    "evidence_reference",
    "narrative_lines",
    "no_gap_sentence",
    "no_roadmap_sentence",
    "render_docx",
    "render_pdf",
    "render_xlsx",
    "roadmap_rows",
    "stage_hex",
]
