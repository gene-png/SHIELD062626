"""Deliverable renderers - turn a capability list into XLSX + PDF bytes.

Master Spec §15 Phase 3: "PDF + XLSX exporters for the deliverable."

XLSX: openpyxl. Header row + one row per capability + a summary row at
the bottom (Total Cost, Estimated Savings).

PDF: ReportLab. Pure Python; no native deps required (unlike WeasyPrint).
Phase 6 polish can revisit visual fidelity, but for v1 the deliverable is
a real, legitimate PDF with a title, summary, table, and savings figure.

Both renderers are pure functions over the data; no DB, no I/O. The
route layer writes the bytes via the existing StorageBackend.
"""

from __future__ import annotations

import io
import logging
from collections.abc import Iterable
from dataclasses import dataclass

from app import export_style
from app.models.capability import CapabilityDisposition, CapabilityItem, CapabilityList

logger = logging.getLogger(__name__)
_LOG = "tech_debt.exporters:"


@dataclass(frozen=True)
class DeliverableContext:
    """Inputs the renderers share. Built once by the route layer."""

    client_legal_name: str
    service_title: str
    cap_list: CapabilityList
    items: list[CapabilityItem]
    total_cost: float
    estimated_savings: float
    savings_cost_known: bool


def _disposition_label(d: CapabilityDisposition | None) -> str:
    if d is None:
        return "Undecided"
    return {
        CapabilityDisposition.KEEP: "Keep",
        CapabilityDisposition.CONSOLIDATE: "Consolidate",
        CapabilityDisposition.CUT: "Cut",
    }[d]


def build_context(
    *,
    client_legal_name: str | None,
    service_title: str,
    cap_list: CapabilityList,
    items: Iterable[CapabilityItem],
) -> DeliverableContext:
    items_list = list(items)
    total_cost = 0.0
    estimated_savings = 0.0
    savings_known = True
    for it in items_list:
        if it.annual_cost_usd is not None:
            total_cost += float(it.annual_cost_usd)
        if it.disposition == CapabilityDisposition.CUT:
            if it.annual_cost_usd is None:
                savings_known = False
            else:
                estimated_savings += float(it.annual_cost_usd)
    return DeliverableContext(
        client_legal_name=client_legal_name or "Client",
        service_title=service_title,
        cap_list=cap_list,
        items=items_list,
        total_cost=total_cost,
        estimated_savings=estimated_savings,
        savings_cost_known=savings_known,
    )


# ---------------------------------------------------------------------------
# Portfolio narrative — computed from the rows, never canned
#
# Every figure below is derived from the capability list. Each sentence states
# the absence of its input explicitly rather than degrading into a reassuring
# claim, which is the defect shape Sprint 10 S3/S4 kept producing: a sentence
# that reads as a finding while resting on data nobody supplied.
# ---------------------------------------------------------------------------

_EMPTY_PORTFOLIO_SENTENCE = (
    "No capabilities are recorded on this list, so there is no portfolio to "
    "summarize, no cost drivers to rank, and no savings to estimate."
)


def _plural(count: int, singular: str, plural: str) -> str:
    return singular if count == 1 else plural


def _disposition_sentence(items: list[CapabilityItem]) -> str:
    counts = {
        d: sum(1 for it in items if it.disposition == d)
        for d in (
            CapabilityDisposition.KEEP,
            CapabilityDisposition.CONSOLIDATE,
            CapabilityDisposition.CUT,
        )
    }
    total = len(items)
    undecided = total - sum(counts.values())
    noun = _plural(total, "capability", "capabilities")
    if undecided == total:
        return (
            f"None of the {total} {noun} carries a disposition yet, so this list "
            "records the inventory only and no keep, consolidate or cut split can "
            "be reported."
        )
    split = (
        f"{counts[CapabilityDisposition.KEEP]} Keep, "
        f"{counts[CapabilityDisposition.CONSOLIDATE]} Consolidate and "
        f"{counts[CapabilityDisposition.CUT]} Cut"
    )
    tail = f", with {undecided} still Undecided." if undecided else "."
    return f"Of the {total} {noun} reviewed, {split}{tail}"


def _cost_sentence(items: list[CapabilityItem], total_cost: float) -> str:
    costed = sorted(
        (it for it in items if it.annual_cost_usd is not None),
        key=lambda it: (-float(it.annual_cost_usd or 0), it.name),
    )
    if not costed:
        return (
            "No annual cost is recorded on any row, so no cost drivers can be "
            "ranked and no savings can be estimated."
        )
    missing = len(items) - len(costed)
    drivers = ", ".join(f"{it.name} (${float(it.annual_cost_usd or 0):,.0f})" for it in costed[:3])
    lead = (
        f"The one row carrying a cost totals ${total_cost:,.0f} a year"
        if len(costed) == 1
        else f"The {len(costed)} rows carrying a cost total ${total_cost:,.0f} a year"
    )
    floor = (
        f", and {missing} of the {len(items)} rows "
        f"{_plural(missing, 'carries', 'carry')} no cost, so that total is a floor"
        if missing
        else ""
    )
    largest = _plural(len(costed[:3]), "the largest is", "the largest are")
    return f"{lead}{floor}; {largest} {drivers}."


def _savings_sentence(ctx: DeliverableContext) -> str:
    cut = [it for it in ctx.items if it.disposition == CapabilityDisposition.CUT]
    if not cut:
        return "No row is marked Cut, so no annual savings are claimed."
    uncosted = sum(1 for it in cut if it.annual_cost_usd is None)
    if uncosted == len(cut):
        return (
            f"{len(cut)} {_plural(len(cut), 'row', 'rows')} marked Cut "
            f"{_plural(len(cut), 'carries', 'carry')} no annual cost, so no savings "
            "figure can be computed from this list."
        )
    rows = f"{len(cut)} {_plural(len(cut), 'row', 'rows')} marked Cut"
    if uncosted:
        return (
            f"Cutting the {rows} removes at least ${ctx.estimated_savings:,.0f} of "
            f"annual spend; {uncosted} Cut {_plural(uncosted, 'row', 'rows')} "
            f"{_plural(uncosted, 'carries', 'carry')} no cost, so the figure is a "
            "lower bound."
        )
    share = (
        f", {round(ctx.estimated_savings / ctx.total_cost * 100)}% of the recorded total"
        if ctx.total_cost > 0
        else ""
    )
    return f"Cutting the {rows} removes ${ctx.estimated_savings:,.0f} of annual spend{share}."


def portfolio_paragraph(ctx: DeliverableContext) -> list[str]:
    """Compute the portfolio narrative sentences for a capability list.

    Sentence 1 counts the dispositions, sentence 2 ranks the top cost drivers,
    sentence 3 frames the savings and preserves the lower-bound caveat when a
    Cut row is missing a cost. Thin data produces an explicit statement of
    absence, never an implied finding.
    """
    if not ctx.items:
        logger.debug("%s portfolio_paragraph on an empty capability list", _LOG)
        return [_EMPTY_PORTFOLIO_SENTENCE]
    sentences = [
        _disposition_sentence(ctx.items),
        _cost_sentence(ctx.items, ctx.total_cost),
        _savings_sentence(ctx),
    ]
    logger.debug(
        "%s portfolio_paragraph items=%d cost=%.0f savings=%.0f cost_known=%s -> %d sentences",
        _LOG,
        len(ctx.items),
        ctx.total_cost,
        ctx.estimated_savings,
        ctx.savings_cost_known,
        len(sentences),
    )
    return sentences


def portfolio_text(ctx: DeliverableContext) -> str:
    """The portfolio narrative as one paragraph of prose."""
    return " ".join(portfolio_paragraph(ctx))


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def render_xlsx(ctx: DeliverableContext) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("openpyxl returned no active worksheet")
    ws.title = "Capability List"

    header = [
        "Name",
        "Vendor",
        "Category",
        "Function",
        "Annual Cost (USD)",
        "Licenses",
        "Disposition",
        "Rationale",
        "Notes",
        "AI Confidence %",
    ]
    ws.append(header)
    header_fill = export_style.xlsx_header_fill()
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for item in ctx.items:
        ws.append(
            [
                item.name,
                item.vendor or "",
                item.category or "",
                item.function or "",
                float(item.annual_cost_usd) if item.annual_cost_usd is not None else "",
                item.license_count if item.license_count is not None else "",
                _disposition_label(item.disposition),
                item.disposition_rationale or "",
                item.notes or "",
                item.confidence_pct if item.confidence_pct is not None else "",
            ]
        )

    # Summary row at the bottom.
    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value="Total annual cost").font = Font(bold=True)
    ws.cell(row=summary_row, column=5, value=ctx.total_cost).number_format = "$#,##0"
    ws.cell(row=summary_row + 1, column=1, value="Estimated annual savings").font = Font(bold=True)
    savings_cell = ws.cell(row=summary_row + 1, column=5, value=ctx.estimated_savings)
    savings_cell.number_format = "$#,##0"
    if not ctx.savings_cost_known:
        ws.cell(
            row=summary_row + 1,
            column=6,
            value="≥ (one or more cut rows missing a cost)",
        ).font = Font(italic=True)

    narrative_row = summary_row + 3
    ws.cell(row=narrative_row, column=1, value="Portfolio summary").font = Font(bold=True)
    ws.cell(row=narrative_row + 1, column=1, value=portfolio_text(ctx)).alignment = Alignment(
        vertical="top", wrap_text=True
    )

    # Reasonable column widths.
    widths = [28, 22, 16, 28, 18, 10, 14, 38, 38, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def render_pdf(ctx: DeliverableContext) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )

    out = io.BytesIO()
    doc = export_style.new_pdf_doc(
        out,
        title=export_style.metadata_title(ctx.service_title, ctx.client_legal_name),
        side_margin_in=export_style.SERVICE_PAGE_MARGIN_IN,
    )
    styles = getSampleStyleSheet()
    h1 = styles["Title"]
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], spaceBefore=14, spaceAfter=6)
    body = styles["BodyText"]

    story: list = []
    story.append(
        Paragraph(export_style.escaped_title(ctx.service_title, ctx.client_legal_name), h1)
    )
    story.append(Paragraph(export_style.escaped_line(ctx.client_legal_name), body))
    story.append(Spacer(1, 0.2 * inch))

    story.append(Paragraph("Summary", h2))
    savings = (
        f"${ctx.estimated_savings:,.0f}"
        if ctx.savings_cost_known
        else f"≥ ${ctx.estimated_savings:,.0f}"
    )
    story.append(
        Paragraph(
            f"Capabilities reviewed: <b>{len(ctx.items)}</b> · "
            f"Total annual cost: <b>${ctx.total_cost:,.0f}</b> · "
            f"Estimated annual savings: <b>{savings}</b>",
            body,
        )
    )
    if not ctx.savings_cost_known:
        story.append(
            Paragraph(
                "Note: at least one row marked <i>Cut</i> is missing an annual cost. "
                "The savings figure is a lower bound.",
                body,
            )
        )

    story.append(Paragraph("<b>Portfolio summary</b>", body))
    story.append(Paragraph(export_style.escaped_line(portfolio_text(ctx)), body))

    story.append(Paragraph("Capability list", h2))

    table_data: list[list] = [["Name", "Vendor", "Category", "Annual cost", "Disposition"]]
    for item in ctx.items:
        cost = f"${float(item.annual_cost_usd):,.0f}" if item.annual_cost_usd is not None else "—"
        table_data.append(
            [
                item.name,
                item.vendor or "",
                item.category or "",
                cost,
                _disposition_label(item.disposition),
            ]
        )

    table = Table(
        table_data,
        colWidths=[2.2 * inch, 1.4 * inch, 1.2 * inch, 1.0 * inch, 1.2 * inch],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(export_style.SURFACE_SUNKEN_HEX)),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(export_style.INK_HEX)),
                ("ALIGN", (3, 1), (3, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(export_style.BORDER_HEX)),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ]
        )
    )
    story.append(table)

    doc.build(story)
    return out.getvalue()


# ---------------------------------------------------------------------------
# DOCX (Work Order C4) - mirrors the PDF content.
# ---------------------------------------------------------------------------


def render_docx(ctx: DeliverableContext) -> bytes:
    from app.docx_export import (
        add_heading,
        add_paragraphs,
        add_table,
        add_title,
        new_document,
        to_bytes,
    )

    doc = new_document(export_style.metadata_title(ctx.service_title, ctx.client_legal_name))
    add_title(doc, ctx.service_title, ctx.client_legal_name)

    savings = (
        f"${ctx.estimated_savings:,.0f}"
        if ctx.savings_cost_known
        else f"≥ ${ctx.estimated_savings:,.0f}"
    )
    add_heading(doc, "Summary")
    lines = [
        f"Capabilities reviewed: {len(ctx.items)}",
        f"Total annual cost: ${ctx.total_cost:,.0f}",
        f"Estimated annual savings: {savings}",
    ]
    if not ctx.savings_cost_known:
        lines.append(
            "Note: at least one row marked Cut is missing an annual cost. "
            "The savings figure is a lower bound."
        )
    add_paragraphs(doc, lines)

    add_heading(doc, "Portfolio summary")
    add_paragraphs(doc, [portfolio_text(ctx)])

    add_heading(doc, "Capability list")
    rows = []
    for item in ctx.items:
        cost = f"${float(item.annual_cost_usd):,.0f}" if item.annual_cost_usd is not None else "—"
        rows.append(
            [
                item.name,
                item.vendor or "",
                item.category or "",
                cost,
                _disposition_label(item.disposition),
            ]
        )
    add_table(doc, ["Name", "Vendor", "Category", "Annual cost", "Disposition"], rows)

    return to_bytes(doc)
