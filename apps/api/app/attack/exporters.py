"""ATT&CK Coverage deliverable renderers - PDF + XLSX.

XLSX sheets:
  - Heatmap Summary: tactic rollup (counts + coverage %, shaded on the shared
                     sequential ramp)
  - Coverage:        per-technique status, curated tool citations, rationale,
                     evidence reference and notes (all 600+ rows)
  - Gaps:            techniques flagged as Gap, ordered by tactic, with a
                     Gap Direction cell stating what the row cites

PDF:
  Executive page with overall coverage %, the citation defensibility stat, the
  methodology disclosure and the per-tactic table, then the Gap list (top 50).

Label discipline (D-035). `routes/attack.py` has Run AI overwrite the tool
columns and the rationale on every unlocked row, so those fields are AI-drafted
unless a consultant edited or locked the row, and no acceptance state is
recorded yet. Every label this module emits therefore states a citation fact and
nothing else: never why a gap exists, never what to do about it.
"""

from __future__ import annotations

import io
import logging
import uuid
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass, field
from typing import TYPE_CHECKING, Any

from app import export_style
from app.attack.analytics import CoverageRollup
from app.attack.catalog import TACTICS, TECHNIQUES, technique_by_id
from app.attack.coverage import CoverageStatus, coverage_label
from app.models.attack_assessment import AttackAssessment, AttackCoverage

if TYPE_CHECKING:
    from reportlab.platypus import TableStyle

logger = logging.getLogger(__name__)
_LOG = "attack.exporters:"

# Printed when `evidence_artifact_id` is genuinely NULL. A dangling pointer
# raises instead, so this sentence never covers a failed lookup.
NO_EVIDENCE_REFERENCE = "No evidence attached"

GAP_DIRECTION_NO_TOOLS = "No detection, prevention, or response tool is cited for this technique"

# One band per step of export_style.GRADED_RAMP_HEX.
COVERAGE_BANDS = 7

METHODOLOGY_NOTE: tuple[str, ...] = (
    "Coverage status, the detection, prevention and response tool citations, and "
    "the rationale on every technique are drafted by Run AI from this client's "
    "capability list. Each cited tool is checked against that list before it is "
    "stored.",
    "A consultant can edit any of these fields, and locking a row keeps a later "
    "Run AI run from overwriting it. An unlocked, unedited row therefore carries "
    "AI-drafted values.",
    "This report states what each row cites. Nothing here records that a "
    "consultant substantiated a citation, so no field here should be read as "
    "verified. Substantiation states arrive in the next batch of work.",
)


@dataclass(frozen=True)
class AttackDeliverableContext:
    client_legal_name: str
    service_title: str
    assessment: AttackAssessment
    coverage: list[AttackCoverage]
    rollup: CoverageRollup
    # artifact id -> filename, resolved by the route's join. Every non-NULL
    # `evidence_artifact_id` on `coverage` must appear here.
    evidence_names: dict[uuid.UUID, str] = field(default_factory=dict)


def build_context(
    *,
    client_legal_name: str | None,
    service_title: str,
    assessment: AttackAssessment,
    coverage: Iterable[AttackCoverage],
    rollup: CoverageRollup,
    evidence_names: Mapping[uuid.UUID, str] | None = None,
) -> AttackDeliverableContext:
    rows = list(coverage)
    names = dict(evidence_names or {})
    unresolved = sorted(
        str(r.evidence_artifact_id)
        for r in rows
        if r.evidence_artifact_id is not None and r.evidence_artifact_id not in names
    )
    if unresolved:
        raise ValueError(
            f"{_LOG} evidence_names is missing {len(unresolved)} cited artifact id(s): "
            f"{unresolved[:5]}. Rendering '{NO_EVIDENCE_REFERENCE}' for a row that "
            f"does cite evidence would misstate the record."
        )
    logger.debug("%s build_context rows=%d evidence_names=%d", _LOG, len(rows), len(names))
    return AttackDeliverableContext(
        client_legal_name=client_legal_name or "Client",
        service_title=service_title,
        assessment=assessment,
        coverage=rows,
        rollup=rollup,
        evidence_names=names,
    )


def _status_or_unscored(value: str | None) -> str:
    if value is None:
        return "Unscored"
    try:
        return coverage_label(CoverageStatus(value))
    except ValueError:
        return "Unknown"


def _tactic_name(tactic_id: str) -> str:
    for t in TACTICS:
        if t.id == tactic_id:
            return t.name
    return tactic_id


# ---------------------------------------------------------------------------
# Curated citations (D-035)
# ---------------------------------------------------------------------------


def _joined_tools(tools: list | None) -> str:
    """Tool names as the deliverable prints them. Blank when none are cited."""
    if not tools:
        return ""
    return ", ".join(str(t).strip() for t in tools if str(t).strip())


def cited_tools(cov: AttackCoverage) -> list[str]:
    """Every distinct tool cited across detection, prevention and response."""
    out: list[str] = []
    for bucket in (cov.detection_tools, cov.prevention_tools, cov.response_tools):
        for name in bucket or []:
            text = str(name).strip()
            if text and text not in out:
                out.append(text)
    return out


def gap_direction(cov: AttackCoverage) -> str:
    """What a Gap row cites. A citation fact, never a cause or a remedy (D-035)."""
    cited = cited_tools(cov)
    if not cited:
        return GAP_DIRECTION_NO_TOOLS
    return f"Cited: {', '.join(cited)} (partial)"


def _evidence_reference(cov: AttackCoverage | None, names: Mapping[uuid.UUID, str]) -> str:
    """The attached artifact's filename, or the NULL sentence. Never a lookup miss."""
    if cov is None or cov.evidence_artifact_id is None:
        return NO_EVIDENCE_REFERENCE
    try:
        return names[cov.evidence_artifact_id]
    except KeyError as exc:
        raise KeyError(
            f"{_LOG} technique {cov.technique_code} cites evidence artifact "
            f"{cov.evidence_artifact_id} which the route did not resolve"
        ) from exc


def citation_counts(coverage: Sequence[AttackCoverage]) -> tuple[int, int]:
    """(rows citing at least one tool, scored rows). Code computes; AI never drafts."""
    scored = [c for c in coverage if c.status is not None]
    citing = sum(1 for c in scored if cited_tools(c))
    logger.debug("%s citation_counts %d/%d", _LOG, citing, len(scored))
    return citing, len(scored)


def defensibility_stat(coverage: Sequence[AttackCoverage]) -> str:
    """The one-line citation stat both the PDF and the DOCX print."""
    citing, scored = citation_counts(coverage)
    return f"{citing} of {scored} scored techniques cite at least one tool"


def coverage_band(coverage_pct: float) -> int:
    """1..COVERAGE_BANDS for a 0-100 percentage. Raises outside that range."""
    if not 0.0 <= coverage_pct <= 100.0:
        raise ValueError(f"{_LOG} coverage_pct must be within 0..100, got {coverage_pct!r}")
    if coverage_pct == 100.0:
        return COVERAGE_BANDS  # the top band is closed at 100
    return 1 + int(coverage_pct / 100.0 * COVERAGE_BANDS)


def coverage_hex(coverage_pct: float) -> str:
    """Fill colour for a coverage percentage on S1's shared sequential ramp."""
    fill = export_style.graded_hex(coverage_band(coverage_pct), COVERAGE_BANDS)
    logger.debug("%s coverage_hex %.1f%% -> %s", _LOG, coverage_pct, fill)
    return fill


def _argb(hex_color: str) -> str:
    """openpyxl wants ARGB. Derived from a brand hex, never hand-copied."""
    return "FF" + hex_color.lstrip("#").upper()


def _coverage_fill(coverage_pct: float) -> Any:
    """Solid fill for a coverage percentage, off S1's shared sequential ramp."""
    from openpyxl.styles import PatternFill

    argb = _argb(coverage_hex(coverage_pct))
    return PatternFill(start_color=argb, end_color=argb, fill_type="solid")


def _coverage_ink(coverage_pct: float) -> str:
    """AA-safe font colour to print on `_coverage_fill(coverage_pct)`."""
    return _argb(export_style.graded_ink_hex(coverage_band(coverage_pct), COVERAGE_BANDS))


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def render_xlsx(ctx: AttackDeliverableContext) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    header_fill = export_style.xlsx_header_fill()
    bold = Font(bold=True)
    italic = Font(italic=True)

    # --- Heatmap Summary ---
    ws = wb.create_sheet("Heatmap Summary")
    ws.append(["Engagement", ctx.client_legal_name])
    ws.append(["Service", ctx.service_title])
    ws.append(["Assessment version", ctx.assessment.version])
    ws.append(["Coverage %", ctx.rollup.coverage_pct])
    ws.append(
        [
            "Scored / Total",
            f"{ctx.rollup.scored_count}/{ctx.rollup.scored_count + ctx.rollup.unscored_count}",
        ]
    )
    for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=1):
        for cell in row:
            cell.font = bold
    ws.append([])
    headers = [
        "Tactic",
        "Name",
        "Techniques",
        "Sub-techniques",
        "Covered",
        "Partial",
        "Gap",
        "N/A",
        "Unscored",
        "Coverage %",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = bold
        cell.fill = header_fill
    pct_col = len(headers)  # "Coverage %" is the last column
    ws.cell(row=4, column=2).fill = _coverage_fill(ctx.rollup.coverage_pct)
    for tc in ctx.rollup.by_tactic:
        ws.append(
            [
                tc.tactic_id,
                tc.tactic_name,
                tc.technique_count,
                tc.sub_technique_count,
                tc.covered,
                tc.partial,
                tc.gap,
                tc.not_applicable,
                tc.unscored,
                tc.coverage_pct,
            ]
        )
        cell = ws.cell(row=ws.max_row, column=pct_col)
        cell.fill = _coverage_fill(tc.coverage_pct)
        cell.font = Font(color=_coverage_ink(tc.coverage_pct))
    widths = [10, 28, 12, 14, 10, 10, 8, 8, 12, 14]
    for w, col in zip(widths, range(1, len(widths) + 1), strict=True):
        ws.column_dimensions[get_column_letter(col)].width = w

    # --- Coverage (per-technique) ---
    ws2 = wb.create_sheet("Coverage")
    headers2 = [
        "Technique",
        "Name",
        "Tactic(s)",
        "Type",
        "Status",
        "Detection tools",
        "Prevention tools",
        "Response tools",
        "Rationale",
        "Evidence reference",
        "Notes",
    ]
    ws2.append(headers2)
    for col in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    cov_by_code = {c.technique_code: c for c in ctx.coverage}
    for tech in TECHNIQUES:
        cov = cov_by_code.get(tech.id)
        tactic_str = ", ".join(_tactic_name(t) for t in tech.tactics)
        ws2.append(
            [
                tech.id,
                tech.name,
                tactic_str,
                "sub" if tech.is_sub_technique else "parent",
                _status_or_unscored(cov.status if cov else None),
                _joined_tools(cov.detection_tools if cov else None),
                _joined_tools(cov.prevention_tools if cov else None),
                _joined_tools(cov.response_tools if cov else None),
                (cov.rationale if cov and cov.rationale else "") or "",
                _evidence_reference(cov, ctx.evidence_names),
                (cov.notes if cov and cov.notes else "") or "",
            ]
        )
    widths2 = [14, 38, 28, 8, 12, 24, 24, 24, 60, 32, 60]
    for w, col in zip(widths2, range(1, len(widths2) + 1), strict=True):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # --- Gaps ---
    ws3 = wb.create_sheet("Gaps")
    headers3 = ["Technique", "Name", "Tactic(s)", "Gap Direction", "Notes"]
    ws3.append(headers3)
    for col in range(1, len(headers3) + 1):
        cell = ws3.cell(row=1, column=col)
        cell.font = bold
        cell.fill = header_fill
    gap_rows = [c for c in ctx.coverage if c.status == CoverageStatus.GAP.value]
    gap_rows.sort(key=lambda c: c.technique_code)
    for cov in gap_rows:
        try:
            tech = technique_by_id(cov.technique_code)
            tactic_str = ", ".join(_tactic_name(t) for t in tech.tactics)
            name = tech.name
        except KeyError:
            tactic_str = ""
            name = cov.technique_code
        ws3.append(
            [
                cov.technique_code,
                name,
                tactic_str,
                gap_direction(cov),
                cov.notes or "",
            ]
        )
    if not gap_rows:
        ws3.append(["—", "No gaps recorded", "", "", ""])
        ws3.cell(row=2, column=2).font = italic
    widths3 = [14, 38, 28, 62, 60]
    for w, col in zip(widths3, range(1, len(widths3) + 1), strict=True):
        ws3.column_dimensions[get_column_letter(col)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def render_docx(ctx: AttackDeliverableContext) -> bytes:
    """Word deliverable mirroring the PDF (Work Order C4)."""
    from app.docx_export import (
        add_heading,
        add_paragraphs,
        add_table,
        add_title,
        new_document,
        to_bytes,
    )

    doc = new_document(export_style.metadata_title(ctx.service_title, ctx.client_legal_name))
    add_title(doc, ctx.service_title, ctx.client_legal_name)

    add_heading(doc, "Coverage summary")
    add_paragraphs(
        doc,
        [
            f"Overall coverage: {ctx.rollup.coverage_pct}%",
            f"Scored: {ctx.rollup.scored_count}/"
            f"{ctx.rollup.scored_count + ctx.rollup.unscored_count}",
            f"Covered {ctx.rollup.covered}, Partial {ctx.rollup.partial}, "
            f"Gap {ctx.rollup.gap}, N/A {ctx.rollup.not_applicable}",
            defensibility_stat(ctx.coverage),
        ],
    )

    add_heading(doc, "Methodology and what this report does not claim")
    add_paragraphs(doc, METHODOLOGY_NOTE)

    add_heading(doc, "Per-tactic rollup")
    add_table(
        doc,
        ["Tactic", "Name", "Covered", "Partial", "Gap", "N/A", "Coverage %"],
        [
            [
                tc.tactic_id,
                tc.tactic_name,
                tc.covered,
                tc.partial,
                tc.gap,
                tc.not_applicable,
                f"{tc.coverage_pct}%",
            ]
            for tc in ctx.rollup.by_tactic
        ],
    )

    gap_rows = [c for c in ctx.coverage if c.status == CoverageStatus.GAP.value]
    gap_rows.sort(key=lambda c: c.technique_code)
    gap_rows = gap_rows[:50]
    add_heading(doc, f"Top remediation gaps ({len(gap_rows)} of {ctx.rollup.gap} shown)")
    if not gap_rows:
        add_paragraphs(doc, ["No techniques flagged as Gap."])
    else:
        rows = []
        for cov in gap_rows:
            try:
                name = technique_by_id(cov.technique_code).name
            except KeyError:
                name = ""
            rows.append([cov.technique_code, name])
        add_table(doc, ["Code", "Technique"], rows)

    return to_bytes(doc)


def render_pdf(ctx: AttackDeliverableContext) -> bytes:
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        Spacer,
        Table,
    )

    out = io.BytesIO()
    doc = export_style.new_pdf_doc(
        out,
        title=export_style.metadata_title(ctx.service_title, ctx.client_legal_name),
        side_margin_in=export_style.SERVICE_PAGE_MARGIN_IN,
    )
    styles = getSampleStyleSheet()
    h1 = styles["Title"]
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], spaceBefore=14, spaceAfter=6)
    body = styles["BodyText"]

    story: list = []
    story.append(
        Paragraph(export_style.escaped_title(ctx.service_title, ctx.client_legal_name), h1)
    )
    story.append(Paragraph(export_style.escaped_line(ctx.client_legal_name), body))
    story.append(Spacer(1, 0.2 * inch))

    story.append(Paragraph("Coverage summary", h2))
    story.append(
        Paragraph(
            f"Overall coverage: <b>{ctx.rollup.coverage_pct}%</b> · "
            f"Scored: <b>{ctx.rollup.scored_count}/"
            f"{ctx.rollup.scored_count + ctx.rollup.unscored_count}</b> · "
            f"Covered <b>{ctx.rollup.covered}</b>, "
            f"Partial <b>{ctx.rollup.partial}</b>, "
            f"Gap <b>{ctx.rollup.gap}</b>, "
            f"N/A <b>{ctx.rollup.not_applicable}</b>",
            body,
        )
    )
    # Computed in Python from the coverage rows, never drafted by a model.
    story.append(Paragraph(export_style.escaped_line(defensibility_stat(ctx.coverage)), body))

    story.append(Paragraph("Methodology and what this report does not claim", h2))
    for line in METHODOLOGY_NOTE:
        story.append(Paragraph(export_style.escaped_line(line), body))

    story.append(Paragraph("Per-tactic rollup", h2))
    tactic_table_data: list[list] = [
        ["Tactic", "Name", "Covered", "Partial", "Gap", "N/A", "Coverage %"]
    ]
    for tc in ctx.rollup.by_tactic:
        tactic_table_data.append(
            [
                tc.tactic_id,
                tc.tactic_name,
                tc.covered,
                tc.partial,
                tc.gap,
                tc.not_applicable,
                f"{tc.coverage_pct}%",
            ]
        )
    tactic_col_widths = [
        0.8 * inch,
        1.9 * inch,
        0.7 * inch,
        0.7 * inch,
        0.6 * inch,
        0.6 * inch,
        0.9 * inch,
    ]
    tactic_table = Table(
        tactic_table_data,
        colWidths=tactic_col_widths,
        repeatRows=1,
    )
    tactic_table.setStyle(_table_style())
    story.append(tactic_table)

    story.append(PageBreak())

    # Top-50 gap list.
    gap_rows = [c for c in ctx.coverage if c.status == CoverageStatus.GAP.value]
    gap_rows.sort(key=lambda c: c.technique_code)
    gap_rows = gap_rows[:50]
    story.append(
        Paragraph(
            f"Top remediation gaps ({len(gap_rows)} of {ctx.rollup.gap} shown)",
            h2,
        )
    )
    if not gap_rows:
        story.append(Paragraph("No techniques flagged as Gap.", body))
    else:
        gap_table_data: list[list] = [["Code", "Technique"]]
        for cov in gap_rows:
            try:
                name = technique_by_id(cov.technique_code).name
            except KeyError:
                name = cov.technique_code
            gap_table_data.append([cov.technique_code, name])
        gap_table = Table(
            gap_table_data,
            colWidths=[1.1 * inch, 4.6 * inch],
            repeatRows=1,
        )
        gap_table.setStyle(_table_style())
        story.append(gap_table)

    doc.build(story)
    return out.getvalue()


def _table_style() -> TableStyle:
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle

    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(export_style.SURFACE_SUNKEN_HEX)),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(export_style.INK_HEX)),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(export_style.BORDER_HEX)),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ]
    )


__all__ = [
    "GAP_DIRECTION_NO_TOOLS",
    "METHODOLOGY_NOTE",
    "NO_EVIDENCE_REFERENCE",
    "AttackDeliverableContext",
    "build_context",
    "cited_tools",
    "citation_counts",
    "coverage_band",
    "coverage_hex",
    "defensibility_stat",
    "gap_direction",
    "render_docx",
    "render_pdf",
    "render_xlsx",
]
