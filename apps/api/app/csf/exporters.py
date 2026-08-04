"""CSF 2.0 deliverable renderers - turn an assessment into XLSX + PDF bytes.

Master Spec §15 Phase 4: each released service produces a PDF + XLSX
deliverable. Reuses the structural pattern from
`app.tech_debt.exporters` (reportlab + openpyxl, pure functions, no
DB/IO). The route layer writes the bytes via the existing
StorageBackend abstraction.

XLSX sheets:
  - "Score Summary": overall + per-function maturity rollup, average-tier cells
                     shaded on the shared sequential ramp
  - "Answers":       per-subcategory tier (shaded), notes and evidence reference
  - "Gap Plan":      prioritized gaps carrying their POA&M annotation

PDF / DOCX:
  Executive summary page with overall maturity + per-function bars, the
  four-tier methodology block, then the top-N gap table and the action plan.

S3 note on language. A CSF POA&M is a Plan of Action and Milestones: the
characterization, owner, deadline, resources and success criteria on a
`CsfGapAction` row are typed by a consultant, so remediation framing is
truthful here in a way it is not for the ATT&CK deliverable (D-035). What still
holds is that every number and every next-step sentence is computed in Python
from the rows — no model drafts prose into this document.
"""

from __future__ import annotations

import io
import logging
import uuid
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass, field
from typing import TYPE_CHECKING, Any

from app import export_style
from app.csf.catalog import FUNCTIONS, SUBCATEGORIES, FunctionCode, Subcategory
from app.csf.gap import Gap, GapAnalysis
from app.csf.maturity import TIER_DEFINITIONS, tier_label
from app.csf.scoring import ScoreResult

if TYPE_CHECKING:
    from reportlab.platypus import TableStyle
from app.models.csf_assessment import CsfAnswer, CsfAssessment
from app.models.csf_profile import CsfGapAction

logger = logging.getLogger(__name__)
_LOG = "csf.exporters:"

# Printed when `evidence_artifact_id` is genuinely NULL. An unresolved pointer
# raises in `build_context`, so this sentence never covers a failed lookup.
NO_EVIDENCE_REFERENCE = "No evidence attached"

# NIST's CSF 2.0 ladder has four rungs; every graded fill in this module is
# `graded_hex(tier, TIER_LEVELS)`.
TIER_LEVELS = len(TIER_DEFINITIONS)

# The methodology block, built from TIER_DEFINITIONS so the report and the
# engine can never disagree. The playbook's five-dimension METHODOLOGY text is
# deliberately NOT reused: it describes a different scoring model.
TIER_MODEL_NOTE: tuple[str, ...] = (
    "Every subcategory in this assessment is scored on NIST's four-tier CSF 2.0 "
    "maturity model. Tier averages, gap sizes, priorities and the next steps "
    "below are computed in code from those scores.",
    *(f"Tier {int(d.tier)} — {d.short_label}: {d.description}" for d in TIER_DEFINITIONS),
)

# Columns the Gap Plan carries after the computed ones, mirroring the
# playbook's Action Plan contract (`playbook_export.py:137-178`).
POAM_FIELDS: tuple[tuple[str, str], ...] = (
    ("Characterization", "characterization"),
    ("Owner", "owner"),
    ("Deadline", "deadline"),
    ("Resources", "resources"),
    ("Success criteria", "success_criteria"),
    ("POA&M ref", "poam_ref"),
)


@dataclass(frozen=True)
class CsfDeliverableContext:
    """Inputs the renderers share. Built once by the route layer."""

    client_legal_name: str
    service_title: str
    assessment: CsfAssessment
    answers: list[CsfAnswer]
    score: ScoreResult
    gap: GapAnalysis
    # subcategory_code -> the consultant's POA&M annotation for that gap.
    actions: dict[str, CsfGapAction] = field(default_factory=dict)
    # artifact id -> filename, resolved by the route's join. Every non-NULL
    # `evidence_artifact_id` on `answers` must appear here.
    evidence_names: dict[uuid.UUID, str] = field(default_factory=dict)


def build_context(
    *,
    client_legal_name: str | None,
    service_title: str,
    assessment: CsfAssessment,
    answers: Iterable[CsfAnswer],
    score: ScoreResult,
    gap: GapAnalysis,
    actions: Mapping[str, CsfGapAction] | None = None,
    evidence_names: Mapping[uuid.UUID, str] | None = None,
) -> CsfDeliverableContext:
    rows = list(answers)
    names = dict(evidence_names or {})
    unresolved = sorted(
        str(a.evidence_artifact_id)
        for a in rows
        if a.evidence_artifact_id is not None and a.evidence_artifact_id not in names
    )
    if unresolved:
        raise ValueError(
            f"{_LOG} evidence_names is missing {len(unresolved)} cited artifact id(s): "
            f"{unresolved[:5]}. Rendering '{NO_EVIDENCE_REFERENCE}' for an answer that "
            f"does cite evidence would misstate the record."
        )
    plan = dict(actions or {})
    logger.debug(
        "%s build_context answers=%d actions=%d evidence_names=%d",
        _LOG,
        len(rows),
        len(plan),
        len(names),
    )
    return CsfDeliverableContext(
        client_legal_name=client_legal_name or "Client",
        service_title=service_title,
        assessment=assessment,
        answers=rows,
        score=score,
        gap=gap,
        actions=plan,
        evidence_names=names,
    )


def _function_name(code: FunctionCode) -> str:
    for f in FUNCTIONS:
        if f.code == code:
            return f.name
    return code.value


def _subcategory_meta(code: str) -> Subcategory | None:
    for s in SUBCATEGORIES:
        if s.code == code:
            return s
    return None


def _fmt_tier(value: float | None) -> str:
    if value is None:
        return "—"
    return f"{value:.2f}"


# ---------------------------------------------------------------------------
# POA&M annotations (Sprint 5 T5 machinery, adopted by the released deliverable)
# ---------------------------------------------------------------------------


def action_text(action: CsfGapAction | None, attribute: str) -> str:
    """One POA&M field as the deliverable prints it. Blank when unset."""
    if action is None:
        return ""
    return getattr(action, attribute) or ""


def effective_priority(gap: Gap, action: CsfGapAction | None) -> str:
    """The consultant's `priority_override` where set, else the computed score.

    Mirrors the playbook's `override or (r.priority or "")` contract. The
    computed side is the weighted gap score, formatted the way the PDF has
    always printed it, so both branches yield a string and the column never
    mixes types.
    """
    override = action_text(action, "priority_override")
    if override:
        logger.debug("%s %s priority override %s wins", _LOG, gap.code, override)
        return override
    return f"{gap.priority_score:.2f}"


def evidence_reference(answer: CsfAnswer | None, names: Mapping[uuid.UUID, str]) -> str:
    """The attached artifact's filename, or the NULL sentence. Never a lookup miss."""
    if answer is None or answer.evidence_artifact_id is None:
        return NO_EVIDENCE_REFERENCE
    try:
        return names[answer.evidence_artifact_id]
    except KeyError as exc:
        raise KeyError(
            f"{_LOG} answer {answer.subcategory_code} cites evidence artifact "
            f"{answer.evidence_artifact_id} which the route did not resolve"
        ) from exc


def _no_gap_steps(ctx: CsfDeliverableContext) -> list[str]:
    """What to say when no scored subcategory sits below target.

    `analyze()` only raises a gap for an ANSWERED subcategory below target, so
    an assessment that scored three of 106 produces zero gaps exactly like one
    that scored all 106 at target. Reading both as an all-clear would let the
    deliverable certify 103 subcategories nobody assessed, so the sentence
    depends on coverage: no finding at all when nothing is scored, the unscored
    count when coverage is partial, and only at full coverage the line that
    tells the client to hold their current controls.
    """
    answered = ctx.score.answered_subcategories
    total = ctx.score.total_subcategories
    target = ctx.gap.target_tier
    label = ctx.gap.target_label
    logger.debug("%s _no_gap_steps coverage=%d/%d", _LOG, answered, total)
    if answered == 0:
        return [
            "No subcategory has been scored, so this report records no maturity finding.",
            f"All {total} subcategories in the NIST CSF 2.0 catalog remain unassessed.",
        ]
    if answered < total:
        return [
            f"No scored subcategory fell below target T{target} ({label}).",
            f"{total - answered} of {total} subcategories are unscored and carry no finding.",
        ]
    return [
        f"No subcategory scored below target T{target} ({label}). Maintain the "
        f"current controls and re-assess on the next cycle."
    ]


def next_steps(ctx: CsfDeliverableContext) -> list[str]:
    """The action plan's next-step sentences, computed from the rows.

    Code computes, the model never drafts: each sentence is a count taken off
    `ctx.gap.gaps` and the POA&M annotations attached to them.
    """
    gaps: Sequence[Gap] = ctx.gap.gaps
    target = ctx.gap.target_tier
    label = ctx.gap.target_label
    if not gaps:
        return _no_gap_steps(ctx)

    total = len(gaps)
    widest = max(g.gap_size for g in gaps)
    widest_count = sum(1 for g in gaps if g.gap_size == widest)
    steps = [
        f"Start with the {widest_count} subcategory gap(s) sitting {widest} tier(s) "
        f"below target T{target} ({label}). They carry the largest lift."
    ]
    for attribute, has_verb, missing_verb in (
        ("owner", "name an owner", "assign"),
        ("deadline", "carry a deadline", "set"),
    ):
        filled = sum(1 for g in gaps if action_text(ctx.actions.get(g.code), attribute))
        if filled == total:
            steps.append(f"All {total} gap(s) in this action plan {has_verb}.")
        else:
            steps.append(
                f"{filled} of {total} gap(s) in this action plan {has_verb}; "
                f"{missing_verb} the remaining {total - filled}."
            )
    logger.debug("%s next_steps: %d sentence(s) over %d gap(s)", _LOG, len(steps), total)
    return steps


def action_plan_rows(ctx: CsfDeliverableContext) -> list[list[str]]:
    """The PDF/DOCX action-plan table body. One row per gap the report lists."""
    rows: list[list[str]] = []
    for g in ctx.gap.gaps:
        action = ctx.actions.get(g.code)
        rows.append(
            [
                g.code,
                g.name,
                effective_priority(g, action),
                action_text(action, "characterization"),
                action_text(action, "owner"),
                action_text(action, "deadline"),
            ]
        )
    return rows


# ---------------------------------------------------------------------------
# Tier shading (S1's shared sequential ramp)
# ---------------------------------------------------------------------------


def average_tier_level(value: float) -> int:
    """Nearest whole tier for a computed average. Raises outside 1..4."""
    if not 1.0 <= value <= float(TIER_LEVELS):
        raise ValueError(f"{_LOG} average tier must be within 1..{TIER_LEVELS}, got {value!r}")
    return int(value + 0.5)


def _argb(hex_color: str) -> str:
    """openpyxl wants ARGB. Derived from a brand hex, never hand-copied."""
    return "FF" + hex_color.lstrip("#").upper()


def tier_fill(tier: int) -> Any:
    """Solid fill for a whole tier, off S1's shared sequential ramp."""
    from openpyxl.styles import PatternFill

    argb = _argb(export_style.graded_hex(tier, TIER_LEVELS))
    return PatternFill(start_color=argb, end_color=argb, fill_type="solid")


def tier_font(tier: int) -> Any:
    """AA-safe font to print on `tier_fill(tier)`."""
    from openpyxl.styles import Font

    return Font(color=_argb(export_style.graded_ink_hex(tier, TIER_LEVELS)))


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def render_xlsx(ctx: CsfDeliverableContext) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # Remove the default blank sheet; we add three named ones below.
    default = wb.active
    if default is not None:
        wb.remove(default)

    header_fill = export_style.xlsx_header_fill()
    bold = Font(bold=True)
    italic = Font(italic=True)

    # --- Sheet 1: Score Summary ---
    ws = wb.create_sheet("Score Summary")
    ws.append(["Engagement", ctx.client_legal_name])
    ws.append(["Service", ctx.service_title])
    ws.append(["Assessment version", ctx.assessment.version])
    ws.append(["Overall maturity", ctx.score.overall_maturity_label])
    ws.append(["Average tier", _fmt_tier(ctx.score.average_tier)])
    ws.append(["Coverage", f"{ctx.score.answered_subcategories}/{ctx.score.total_subcategories}"])
    for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=1):
        for cell in row:
            cell.font = bold
    ws.append([])
    ws.append(["Function", "Name", "Answered", "Total", "Coverage %", "Average tier"])
    for col_idx in range(1, 7):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.font = bold
        cell.fill = header_fill
    for fs in ctx.score.by_function:
        ws.append(
            [
                fs.function.value,
                fs.function_name,
                fs.answered_count,
                fs.subcategory_count,
                fs.coverage_pct,
                _fmt_tier(fs.average_tier),
            ]
        )
        # Heatmap: the per-function average tier reads as a magnitude.
        if fs.average_tier is not None:
            level = average_tier_level(fs.average_tier)
            avg_cell = ws.cell(row=ws.max_row, column=6)
            avg_cell.fill = tier_fill(level)
            avg_cell.font = tier_font(level)
    for w, col in zip([10, 28, 12, 10, 14, 16], range(1, 7), strict=True):
        ws.column_dimensions[get_column_letter(col)].width = w

    # --- Sheet 2: Answers ---
    ws2 = wb.create_sheet("Answers")
    headers = [
        "Subcategory",
        "Function",
        "Category",
        "Name",
        "Outcome",
        "Tier",
        "Tier label",
        "Notes",
        "Evidence",
    ]
    ws2.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    answers_by_code = {a.subcategory_code: a for a in ctx.answers}
    # Iterate the canonical catalog order so missing answers still
    # render as blank rows (the assessor sees what wasn't scored).
    for sc in SUBCATEGORIES:
        ans = answers_by_code.get(sc.code)
        tier = ans.maturity_tier if ans else None
        notes = ans.notes if ans else None
        ws2.append(
            [
                sc.code,
                sc.function.value,
                sc.category,
                sc.name,
                sc.outcome,
                tier if tier is not None else "",
                tier_label(tier) if tier is not None else "Unscored",
                notes or "",
                evidence_reference(ans, ctx.evidence_names),
            ]
        )
        # Heatmap: an unscored row stays unfilled, so blank reads as blank.
        if tier is not None:
            tier_cell = ws2.cell(row=ws2.max_row, column=6)
            tier_cell.fill = tier_fill(int(tier))
            tier_cell.font = tier_font(int(tier))
    for w, col in zip([14, 10, 10, 32, 60, 8, 16, 60, 34], range(1, 10), strict=True):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # --- Sheet 3: Gap Plan ---
    ws3 = wb.create_sheet("Gap Plan")
    headers3 = [
        "Subcategory",
        "Function",
        "Category",
        "Name",
        "Current tier",
        "Target tier",
        "Gap size",
        "Priority",
        *(label for label, _ in POAM_FIELDS),
        "Notes",
    ]
    ws3.append(headers3)
    for col in range(1, len(headers3) + 1):
        cell = ws3.cell(row=1, column=col)
        cell.font = bold
        cell.fill = header_fill
    for g in ctx.gap.gaps:
        action = ctx.actions.get(g.code)
        ws3.append(
            [
                g.code,
                g.function.value,
                g.category,
                g.name,
                g.current_tier,
                g.target_tier,
                g.gap_size,
                effective_priority(g, action),
                *(action_text(action, attribute) for _, attribute in POAM_FIELDS),
                g.notes or "",
            ]
        )
    if not ctx.gap.gaps:
        ws3.append(
            [
                "—",
                "",
                "",
                "No gaps at target tier",
                "",
                ctx.gap.target_tier,
                0,
                "0.00",
                *("" for _ in POAM_FIELDS),
                "",
            ]
        )
        ws3.cell(row=2, column=4).font = italic
    widths3 = [14, 10, 10, 32, 14, 14, 12, 12, 18, 24, 14, 30, 40, 18, 50]
    for w, col in zip(widths3, range(1, len(headers3) + 1), strict=True):
        ws3.column_dimensions[get_column_letter(col)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


_ACTION_PLAN_HEADERS: tuple[str, ...] = (
    "Code",
    "Subcategory",
    "Priority",
    "Characterization",
    "Owner",
    "Deadline",
)


def _action_plan_heading(ctx: CsfDeliverableContext) -> str:
    """Names the truncation outright — the plan lists the gaps the report shows."""
    return f"Action plan ({len(ctx.gap.gaps)} of {ctx.gap.total_gap_count} gaps shown)"


def render_docx(ctx: CsfDeliverableContext) -> bytes:
    """Word deliverable mirroring the PDF (Work Order C4)."""
    from app.docx_export import (
        add_heading,
        add_paragraphs,
        add_table,
        add_title,
        new_document,
        to_bytes,
    )

    doc = new_document(export_style.metadata_title(ctx.service_title, ctx.client_legal_name))
    add_title(doc, ctx.service_title, ctx.client_legal_name)

    add_heading(doc, "Maturity summary")
    add_paragraphs(
        doc,
        [
            f"Overall maturity: {ctx.score.overall_maturity_label}",
            f"Average tier: {_fmt_tier(ctx.score.average_tier)}",
            f"Coverage: {ctx.score.answered_subcategories}/"
            f"{ctx.score.total_subcategories} ({ctx.score.coverage_pct}%)",
        ],
    )

    add_heading(doc, "How these tiers are scored")
    add_paragraphs(doc, list(TIER_MODEL_NOTE))

    add_heading(doc, "Per-function rollup")
    add_table(
        doc,
        ["Function", "Name", "Average tier", "Coverage"],
        [
            [
                fs.function.value,
                fs.function_name,
                _fmt_tier(fs.average_tier),
                f"{fs.answered_count}/{fs.subcategory_count} ({fs.coverage_pct}%)",
            ]
            for fs in ctx.score.by_function
        ],
    )

    add_heading(doc, f"Top remediation gaps (target T{ctx.gap.target_tier})")
    if not ctx.gap.gaps:
        add_paragraphs(
            doc,
            [f"No gaps at target tier {ctx.gap.target_tier} " f"({ctx.gap.target_label})."],
        )
    else:
        add_table(
            doc,
            ["Code", "Function", "Subcategory", "Current → Target", "Priority"],
            [
                [
                    g.code,
                    g.function.value,
                    g.name,
                    f"T{g.current_tier} → T{g.target_tier}",
                    f"{g.priority_score:.2f}",
                ]
                for g in ctx.gap.gaps
            ],
        )

    add_heading(doc, _action_plan_heading(ctx))
    add_paragraphs(doc, next_steps(ctx))
    rows = action_plan_rows(ctx)
    if rows:
        add_table(doc, list(_ACTION_PLAN_HEADERS), rows)

    return to_bytes(doc)


def render_pdf(ctx: CsfDeliverableContext) -> bytes:
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        Spacer,
        Table,
    )

    out = io.BytesIO()
    doc = export_style.new_pdf_doc(
        out,
        title=export_style.metadata_title(ctx.service_title, ctx.client_legal_name),
        side_margin_in=export_style.SERVICE_PAGE_MARGIN_IN,
    )
    styles = getSampleStyleSheet()
    h1 = styles["Title"]
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], spaceBefore=14, spaceAfter=6)
    body = styles["BodyText"]

    story: list = []
    story.append(
        Paragraph(export_style.escaped_title(ctx.service_title, ctx.client_legal_name), h1)
    )
    story.append(Paragraph(export_style.escaped_line(ctx.client_legal_name), body))
    story.append(Spacer(1, 0.2 * inch))

    story.append(Paragraph("Maturity summary", h2))
    story.append(
        Paragraph(
            f"Overall maturity: <b>{ctx.score.overall_maturity_label}</b> · "
            f"Average tier: <b>{_fmt_tier(ctx.score.average_tier)}</b> · "
            f"Coverage: <b>{ctx.score.answered_subcategories}/"
            f"{ctx.score.total_subcategories}</b> "
            f"({ctx.score.coverage_pct}%)",
            body,
        )
    )

    story.append(Paragraph("How these tiers are scored", h2))
    for line in TIER_MODEL_NOTE:
        story.append(Paragraph(export_style.escaped_line(line), body))

    story.append(Paragraph("Per-function rollup", h2))
    fn_table_data: list[list] = [["Function", "Name", "Average tier", "Coverage"]]
    for fs in ctx.score.by_function:
        fn_table_data.append(
            [
                fs.function.value,
                fs.function_name,
                _fmt_tier(fs.average_tier),
                f"{fs.answered_count}/{fs.subcategory_count} ({fs.coverage_pct}%)",
            ]
        )
    fn_table = Table(
        fn_table_data,
        colWidths=[0.8 * inch, 2.2 * inch, 1.4 * inch, 2.0 * inch],
        repeatRows=1,
    )
    fn_table.setStyle(_table_style())
    story.append(fn_table)

    story.append(PageBreak())

    story.append(Paragraph(f"Top remediation gaps (target T{ctx.gap.target_tier})", h2))
    if not ctx.gap.gaps:
        story.append(
            Paragraph(
                f"No gaps at target tier {ctx.gap.target_tier} " f"({ctx.gap.target_label}).",
                body,
            )
        )
    else:
        gap_table_data: list[list] = [
            ["Code", "Function", "Subcategory", "Current → Target", "Priority"]
        ]
        for g in ctx.gap.gaps:
            gap_table_data.append(
                [
                    g.code,
                    g.function.value,
                    g.name,
                    f"T{g.current_tier} → T{g.target_tier}",
                    f"{g.priority_score:.2f}",
                ]
            )
        gap_table = Table(
            gap_table_data,
            colWidths=[0.9 * inch, 0.8 * inch, 3.0 * inch, 1.4 * inch, 0.9 * inch],
            repeatRows=1,
        )
        gap_table.setStyle(_table_style())
        story.append(gap_table)

    story.append(Paragraph(export_style.escaped_line(_action_plan_heading(ctx)), h2))
    for step in next_steps(ctx):
        story.append(Paragraph(export_style.escaped_line(f"• {step}"), body))
    plan_rows = action_plan_rows(ctx)
    if plan_rows:
        plan_table_data: list[list] = [list(_ACTION_PLAN_HEADERS)]
        plan_table_data.extend(plan_rows)
        plan_table = Table(
            plan_table_data,
            colWidths=[0.85 * inch, 2.0 * inch, 0.7 * inch, 1.2 * inch, 1.4 * inch, 1.05 * inch],
            repeatRows=1,
        )
        plan_table.setStyle(_table_style())
        story.append(plan_table)

    doc.build(story)
    return out.getvalue()


def _table_style() -> TableStyle:
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle

    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(export_style.SURFACE_SUNKEN_HEX)),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(export_style.INK_HEX)),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(export_style.BORDER_HEX)),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ]
    )


__all__ = [
    "NO_EVIDENCE_REFERENCE",
    "POAM_FIELDS",
    "TIER_LEVELS",
    "TIER_MODEL_NOTE",
    "CsfDeliverableContext",
    "action_plan_rows",
    "action_text",
    "average_tier_level",
    "build_context",
    "effective_priority",
    "evidence_reference",
    "next_steps",
    "render_docx",
    "render_pdf",
    "render_xlsx",
    "tier_fill",
    "tier_font",
]
