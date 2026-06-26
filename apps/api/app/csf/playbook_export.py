"""CSF full-Playbook XLSX export (Work Order D4).

A functional workbook over the data the engine already computes: an Enterprise
Profile sheet (the weighted-floor roll-up per subcategory) plus one sheet per
tier with the five dimension scores and the code-computed total/level/cap. The
methodology doc's exact cell styling can layer on top later; the data + structure
are here.
"""

from __future__ import annotations

import io
from collections.abc import Mapping, Sequence
from typing import Any


def _autofit(ws: Any) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(60, max(10, width + 2))


def render_xlsx(
    *,
    client_name: str,
    version: int,
    enterprise_rows: Sequence[Any],
    tier_profiles: Mapping[str, Sequence[Any]],
) -> bytes:
    """`enterprise_rows` are EnterpriseSubcategory-like; `tier_profiles` maps a
    tier name to its CsfDimensionScoreResponse-like rows (total/level computed)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    head_fill = PatternFill(start_color="FFEEF2F7", end_color="FFEEF2F7", fill_type="solid")

    def _header(ws: Any, cols: list[str]) -> None:
        ws.append(cols)
        for i in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=i)
            cell.font = Font(bold=True)
            cell.fill = head_fill

    ws = wb.active
    ws.title = "Enterprise Profile"
    _header(
        ws,
        [
            "Subcategory", "Function", "Outcome", "High", "Mod", "Low",
            "Enterprise", "Rule", "Target", "Gap", "Priority",
        ],
    )
    for r in enterprise_rows:
        levels = r.tier_levels
        ws.append(
            [
                r.subcategory_code,
                r.function,
                r.name,
                levels.get("high", ""),
                levels.get("moderate", ""),
                levels.get("low", ""),
                f"L{r.enterprise_level}",
                f"#{r.rollup_rule}",
                f"L{r.target_level}" if r.target_level else "",
                "Yes" if r.gap else "",
                r.priority or "",
            ]
        )
    _autofit(ws)

    for tier in ("high", "moderate", "low"):
        rows = tier_profiles.get(tier)
        if not rows:
            continue
        ts = wb.create_sheet(tier.title())
        _header(
            ts,
            [
                "Subcategory", "Governance", "Policy & Process", "Implementation",
                "Monitoring & Measurement", "Continuous Improvement", "Total",
                "Level", "Evidence capped", "In scope", "Target",
            ],
        )
        for row in rows:
            ts.append(
                [
                    row.subcategory_code,
                    row.governance,
                    row.policy,
                    row.implementation,
                    row.monitoring,
                    row.improvement,
                    row.total,
                    f"L{row.level}",
                    "Yes" if row.evidence_capped else "",
                    "Yes" if row.in_scope else "No",
                    f"L{row.target_level}" if row.target_level else "",
                ]
            )
        _autofit(ts)

    # A small cover note so the export is self-describing.
    cover = wb.create_sheet("About", 0)
    cover.append(["SHIELD by Kentro — CSF 2.0 Full Playbook"])
    cover.append([f"Client: {client_name}"])
    cover.append([f"Working profile version: {version}"])
    cover.append([])
    cover.append([
        "Levels are code-computed: total = sum of the five dimensions (0-10); "
        "L1 0-2, L2 3-5, L3 6-7, L4 8-9, L5 10. Enterprise = weighted-floor "
        "roll-up across the tiers in use."
    ])
    cover["A1"].font = Font(bold=True, size=14)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# Narrative report (PDF + Word) — executive summary + function + gap tables
# ---------------------------------------------------------------------------

_PRIORITY_ORDER = {"P1": 0, "P2": 1, "P3": 2}


def _summary_lines(enterprise_rows: Sequence[Any]) -> list[str]:
    total = len(enterprise_rows)
    gaps = sum(1 for r in enterprise_rows if r.gap)
    p1 = sum(1 for r in enterprise_rows if r.priority == "P1")
    p2 = sum(1 for r in enterprise_rows if r.priority == "P2")
    p3 = sum(1 for r in enterprise_rows if r.priority == "P3")
    return [
        f"Subcategories assessed: {total}",
        f"Gaps vs target: {gaps}",
        f"Priority — P1 {p1}, P2 {p2}, P3 {p3}",
    ]


def _function_rows(enterprise_rows: Sequence[Any]) -> list[list[str]]:
    by_fn: dict[str, dict[str, int]] = {}
    for r in enterprise_rows:
        fn = r.function or "—"
        b = by_fn.setdefault(fn, {"count": 0, "gaps": 0, "level_sum": 0})
        b["count"] += 1
        b["gaps"] += 1 if r.gap else 0
        b["level_sum"] += r.enterprise_level
    out: list[list[str]] = []
    for fn in sorted(by_fn):
        b = by_fn[fn]
        avg = round(b["level_sum"] / b["count"], 1) if b["count"] else 0
        out.append([fn, str(b["count"]), f"L{avg}", str(b["gaps"])])
    return out


def _gap_rows(enterprise_rows: Sequence[Any]) -> list[list[str]]:
    gaps = [r for r in enterprise_rows if r.gap]
    gaps.sort(key=lambda r: (_PRIORITY_ORDER.get(r.priority or "P3", 3), r.subcategory_code))
    return [
        [
            r.subcategory_code,
            r.name,
            f"L{r.enterprise_level}",
            f"L{r.target_level}" if r.target_level else "—",
            r.priority or "",
        ]
        for r in gaps
    ]


def render_pdf(
    *, client_name: str, version: int, enterprise_rows: Sequence[Any]
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out, pagesize=letter, leftMargin=0.7 * inch, rightMargin=0.7 * inch,
        topMargin=0.7 * inch, bottomMargin=0.7 * inch,
        title=f"CSF 2.0 Playbook — {client_name}", author="SHIELD by Kentro",
    )
    styles = getSampleStyleSheet()
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], spaceBefore=14, spaceAfter=6)

    def _grid(data: list[list[str]], widths: list[float]) -> Table:
        t = Table(data, colWidths=widths, repeatRows=1)
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2f7")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d6dae3")),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        return t

    story: list = [
        Paragraph(f"NIST CSF 2.0 Full Playbook (v{version})", styles["Title"]),
        Paragraph(client_name, styles["BodyText"]),
        Spacer(1, 0.2 * inch),
        Paragraph("Executive summary", h2),
    ]
    for line in _summary_lines(enterprise_rows):
        story.append(Paragraph(line, styles["BodyText"]))

    story.append(Paragraph("By function", h2))
    story.append(
        _grid(
            [["Function", "Subcategories", "Avg level", "Gaps"], *_function_rows(enterprise_rows)],
            [2.6 * inch, 1.4 * inch, 1.2 * inch, 1.0 * inch],
        )
    )

    story.append(Paragraph("Prioritized gaps", h2))
    gap_rows = _gap_rows(enterprise_rows)
    if gap_rows:
        story.append(
            _grid(
                [["Subcategory", "Outcome", "Current", "Target", "Priority"], *gap_rows],
                [1.0 * inch, 3.0 * inch, 0.8 * inch, 0.8 * inch, 0.8 * inch],
            )
        )
    else:
        story.append(Paragraph("No gaps — every in-scope subcategory meets target.", styles["BodyText"]))

    doc.build(story)
    return out.getvalue()


def render_docx(
    *, client_name: str, version: int, enterprise_rows: Sequence[Any]
) -> bytes:
    from app.docx_export import (
        add_heading,
        add_paragraphs,
        add_table,
        add_title,
        new_document,
        to_bytes,
    )

    doc = new_document(f"CSF 2.0 Playbook — {client_name}")
    add_title(doc, f"NIST CSF 2.0 Full Playbook (v{version})", client_name)

    add_heading(doc, "Executive summary")
    add_paragraphs(doc, _summary_lines(enterprise_rows))

    add_heading(doc, "By function")
    add_table(doc, ["Function", "Subcategories", "Avg level", "Gaps"], _function_rows(enterprise_rows))

    add_heading(doc, "Prioritized gaps")
    gap_rows = _gap_rows(enterprise_rows)
    if gap_rows:
        add_table(
            doc,
            ["Subcategory", "Outcome", "Current", "Target", "Priority"],
            gap_rows,
        )
    else:
        add_paragraphs(doc, ["No gaps — every in-scope subcategory meets target."])

    return to_bytes(doc)
